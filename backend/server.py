from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query, UploadFile, File, WebSocket, WebSocketDisconnect
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta, date
import jwt
import bcrypt
from fastapi.responses import StreamingResponse
import io
import csv
import json
import asyncio
import secrets
import resend
import httpx
import base64
import hashlib
import time
from collections import defaultdict
from groq import Groq
from referral_service import ReferralService
from referral_endpoints import create_referral_endpoints, create_admin_referral_endpoints

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL') or os.environ.get('MONGODB_URI') or os.environ.get('DATABASE_URL')
if not mongo_url:
    raise RuntimeError('MongoDB connection URL is required. Set MONGO_URL, MONGODB_URI, or DATABASE_URL.')

client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'trade_ledger')]

# Initialize Referral Service
referral_service = ReferralService(db)

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'trading_journal_secret')
JWT_ALGORITHM = os.environ.get('JWT_ALGORITHM', 'HS256')
JWT_EXPIRATION_HOURS = int(os.environ.get('JWT_EXPIRATION_HOURS', 24))

# Keys
METAAPI_TOKEN = os.environ.get('METAAPI_TOKEN', '')
RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
TWELVE_DATA_API_KEY = os.environ.get('TWELVE_DATA_API_KEY', '')
FINNHUB_API_KEY = os.environ.get('FINNHUB_API_KEY', 'd7bdvvhr01qgc9t72sc0d7bdvvhr01qgc9t72scg')
ALPHA_VANTAGE_API_KEY = os.environ.get('ALPHA_VANTAGE_API_KEY', 'PGLKVQ8VQQ8QRU61')

# Admin credentials
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'admin@tradeledger.com')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'TradeLedger@Admin2024')

# Frontend URL for password reset links
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'http://localhost:3000')

# Initialize Resend
if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

# Initialize Groq (keeping for fallback, but will use xAI Grok primarily)
GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')
groq_client = None

# Initialize xAI for both TTS and Grok LLM
XAI_API_KEY = os.environ.get('XAI_API_KEY', '')
xai_grok_enabled = bool(XAI_API_KEY)

# AI Insights cache: hash(trade_data+question) -> {text, audio_base64, timestamp}
ai_insights_cache = {}
AI_CACHE_TTL = 300  # 5 minutes

# News cache (in-memory)
news_cache = {
    'data': None,
    'timestamp': None
}
NEWS_CACHE_TTL = 60  # 60 seconds

calendar_cache = {
    'data': None,
    'timestamp': None
}
CALENDAR_CACHE_TTL = 60  # 60 seconds

quotes_cache = {
    'data': None,
    'timestamp': None
}
QUOTES_CACHE_TTL = 5  # 5 seconds (market data needs frequent updates)
if GROQ_API_KEY:
    groq_client = Groq(api_key=GROQ_API_KEY)

async def generate_ai_insights_with_xai_grok(prompt: str, system_prompt: str):
    """Use xAI Grok API for AI insights generation. Returns text response."""
    if not XAI_API_KEY:
        logging.error("XAI_API_KEY not configured")
        raise Exception("xAI Grok API not configured")
    
    try:
        grok_start = time.time()
        
        # xAI Grok uses /v1/responses endpoint (from xAI console documentation)
        url = "https://api.x.ai/v1/responses"
        headers = {
            "Authorization": f"Bearer {XAI_API_KEY}",
            "Content-Type": "application/json",
        }
        
        # Combine system and user prompts into single input
        full_input = f"{system_prompt}\n\n{prompt}"
        
        payload = {
            "model": "grok-beta",  # Using stable grok-beta model
            "max_output_tokens": 4096,
            "stream": False,  # Non-streaming mode
            "input": full_input  # xAI uses "input" field, not "messages"
        }
        
        logging.info("Calling xAI Grok API for insights...")
        logging.debug(f"xAI API URL: {url}")
        logging.debug(f"xAI Model: {payload['model']}")
        logging.debug(f"Input length: {len(full_input)} chars")
        
        async with httpx.AsyncClient(timeout=120) as client:
            res = await client.post(url, headers=headers, json=payload)
        
        logging.info(f"xAI Grok API response status: {res.status_code}")
        
        if res.status_code == 200:
            data = res.json()
            logging.info(f"xAI API response keys: {list(data.keys())}")
            
            # xAI response format: check "output" field first (most likely)
            insight_text = data.get('output')
            
            # Fallback: check other possible fields
            if not insight_text:
                for field in ['response', 'text', 'content', 'message']:
                    if field in data:
                        insight_text = data[field]
                        logging.info(f"Found response in field: {field}")
                        break
            
            # Check for choices array (OpenAI-compatible format)
            if not insight_text and 'choices' in data and len(data['choices']) > 0:
                choice = data['choices'][0]
                if isinstance(choice, dict):
                    insight_text = choice.get('message', {}).get('content') or choice.get('text')
                    if insight_text:
                        logging.info("Found response in choices[0]")
            
            if not insight_text:
                logging.error(f"Could not find text in xAI response. Full response: {json.dumps(data, indent=2)}")
                raise Exception(f"Empty or unrecognized response format. Response keys: {list(data.keys())}")
            
            grok_elapsed = time.time() - grok_start
            logging.info(f"xAI Grok latency: {grok_elapsed:.2f}s, Response length: {len(insight_text)} chars")
            return insight_text
        else:
            error_body = res.text
            logging.error(f"xAI API failed with status {res.status_code}")
            logging.error(f"Response body: {error_body[:1000]}")
            
            # Try to parse error message
            try:
                error_data = res.json()
                error_msg = error_data.get('error', {}).get('message', error_body)
            except:
                error_msg = error_body
            
            raise Exception(f"xAI API error ({res.status_code}): {error_msg[:300]}")
            
    except Exception as e:
        logging.error(f"xAI Grok exception: {str(e)}", exc_info=True)
        raise

async def generate_tts(text):
    """Convert text to speech using xAI TTS API. Returns mp3 bytes or None."""
    if not XAI_API_KEY:
        logging.warning("XAI_API_KEY not configured, skipping TTS")
        return None
    
    # Truncate text to avoid TTS API limits (keep first ~2000 chars)
    tts_text = text[:2000] if len(text) > 2000 else text
    
    try:
        tts_start = time.time()
        
        # xAI TTS endpoint
        url = "https://api.x.ai/v1/tts"
        headers = {
            "Authorization": f"Bearer {XAI_API_KEY}",
            "Content-Type": "application/json",
        }
        payload = {
            "text": tts_text,
            "voice": "Rex",
            "language": "en"
        }
        
        logging.info("Generating TTS audio with xAI...")
        async with httpx.AsyncClient(timeout=30) as client:
            res = await client.post(url, headers=headers, json=payload)
        
        if res.status_code == 200:
            tts_elapsed = time.time() - tts_start
            logging.info(f"TTS success: {tts_elapsed:.2f}s, {len(res.content)} bytes")
            return res.content
        else:
            logging.error(f"TTS failed ({res.status_code}): {res.text[:300]}")
            return None
            
    except Exception as e:
        logging.error(f"TTS exception: {str(e)}", exc_info=True)
        return None


# Create the main app
app = FastAPI(title="TradeLedger API", version="2.1.0")

# Create routers
api_router = APIRouter(prefix="/api")
admin_router = APIRouter(prefix="/api/admin")

# Security
security = HTTPBearer()

# ============ MODELS ============

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    name: str
    referral_code: Optional[str] = None  # Referral code from link

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    created_at: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class AdminTokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    is_admin: bool = True

class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

class MT5AccountCreate(BaseModel):
    name: str
    login: str
    password: str = ""
    server: str
    platform: str = "mt5"
    metaapi_account_id: Optional[str] = None
    currency: str = "USD"

class MT5AccountUpdate(BaseModel):
    currency: Optional[str] = None
    metaapi_account_id: Optional[str] = None

class MT5AccountResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    name: str
    login: str
    server: str
    platform: str
    metaapi_account_id: Optional[str] = None
    currency: str = "USD"
    is_connected: bool = False
    last_sync: Optional[str] = None
    balance: Optional[float] = None
    equity: Optional[float] = None
    created_at: str

class TradeCreate(BaseModel):
    instrument: str
    position: str
    entry_price: float
    exit_price: Optional[float] = None
    quantity: float
    entry_date: str
    exit_date: Optional[str] = None
    notes: Optional[str] = ""
    status: str = "open"
    stop_loss: Optional[float] = None
    take_profit: Optional[float] = None
    commission: Optional[float] = 0
    swap: Optional[float] = 0
    mt5_ticket: Optional[str] = None
    mt5_account_id: Optional[str] = None
    currency: str = "USD"
    # Trading journal fields
    pre_trade_analysis: Optional[str] = ""
    post_trade_review: Optional[str] = ""
    emotions: Optional[str] = ""
    lessons: Optional[str] = ""
    tags: Optional[str] = ""
    rating: Optional[int] = None
    risk_reward: Optional[int] = None
    # Execution checklist
    check_htf: Optional[bool] = False
    check_risk: Optional[bool] = False
    check_plan: Optional[bool] = False
    check_levels: Optional[bool] = False
    check_news: Optional[bool] = False
    # Screenshots
    screenshots: Optional[List[str]] = []

class TradeUpdate(BaseModel):
    instrument: Optional[str] = None
    position: Optional[str] = None
    entry_price: Optional[float] = None
    exit_price: Optional[float] = None
    quantity: Optional[float] = None
    entry_date: Optional[str] = None
    exit_date: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None
    stop_loss: Optional[float] = None
    take_profit: Optional[float] = None
    currency: Optional[str] = None
    # Trading journal fields
    pre_trade_analysis: Optional[str] = None
    post_trade_review: Optional[str] = None
    emotions: Optional[str] = None
    lessons: Optional[str] = None
    tags: Optional[str] = None
    rating: Optional[int] = None
    risk_reward: Optional[int] = None
    # Execution checklist
    check_htf: Optional[bool] = None
    check_risk: Optional[bool] = None
    check_plan: Optional[bool] = None
    check_levels: Optional[bool] = None
    check_news: Optional[bool] = None
    # Screenshots
    screenshots: Optional[List[str]] = None

class TradeResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    instrument: str
    position: str
    entry_price: float
    exit_price: Optional[float] = None
    quantity: float
    entry_date: str
    exit_date: Optional[str] = None
    notes: str = ""
    status: str
    stop_loss: Optional[float] = None
    take_profit: Optional[float] = None
    commission: float = 0
    swap: float = 0
    pnl: Optional[float] = None
    pnl_percentage: Optional[float] = None
    mt5_ticket: Optional[str] = None
    mt5_account_id: Optional[str] = None
    currency: str = "USD"
    created_at: str
    # Trading journal fields
    pre_trade_analysis: str = ""
    post_trade_review: str = ""
    emotions: str = ""
    lessons: str = ""
    tags: str = ""
    rating: Optional[int] = None
    risk_reward: Optional[int] = None
    # Execution checklist
    check_htf: bool = False
    check_risk: bool = False
    check_plan: bool = False
    check_levels: bool = False
    check_news: bool = False
    # Screenshots
    screenshots: List[str] = []

class CandleResponse(BaseModel):
    time: int
    open: float
    high: float
    low: float
    close: float
    volume: float


class AIInsightRequest(BaseModel):
    question: Optional[str] = None


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_access_token(user_id: str, email: str, is_admin: bool = False) -> str:
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {"sub": user_id, "email": email, "exp": expire, "is_admin": is_admin}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_admin_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        is_admin = payload.get("is_admin", False)
        if not is_admin:
            raise HTTPException(status_code=403, detail="Admin access required")
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============ SUBSCRIPTION / TRIAL HELPERS ============

TRIAL_DURATION_DAYS = 14

async def check_subscription_status(user: dict) -> bool:
    """Check if user has active or trial subscription"""
    status = user.get('subscription_status')
    if status in ('active', 'trial'):
        end_date = user.get('subscription_end_date')
        if end_date:
            if isinstance(end_date, str):
                end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            if end_date > datetime.now(timezone.utc):
                return True
    return False

async def assign_trial(user_id: str):
    """Assign free trial to new user"""
    trial_end = datetime.now(timezone.utc) + timedelta(days=TRIAL_DURATION_DAYS)
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "subscription_status": "trial",
            "subscription_plan": None,
            "subscription_start_date": datetime.now(timezone.utc).isoformat(),
            "subscription_end_date": trial_end.isoformat(),
            "role": "user",
            "status": "active"
        }}
    )

async def require_active_subscription(user: dict = Depends(get_current_user)):
    """Require active subscription to access feature"""
    is_active = await check_subscription_status(user)
    if not is_active:
        raise HTTPException(
            status_code=403,
            detail="Active subscription required to access this feature"
        )
    return user

# ============ PNL CALCULATION ============

# Contract size multipliers for different instruments
# This converts price difference to actual P&L in USD
INSTRUMENT_MULTIPLIERS = {
    # Commodities - Gold/Silver (price per oz, 100 oz per lot)
    'XAU/USD': 100,     # 1 lot = 100 oz, $1 move = $100/lot
    'XAUUSD': 100,
    'Gold': 100,
    'XAG/USD': 5000,    # 1 lot = 5000 oz, $1 move = $5000/lot
    'XAGUSD': 5000,
    'Silver': 5000,
    
    # Forex pairs - Standard lot = 100,000 units
    # For pairs with USD as quote: pip value = $10 per lot per pip (0.0001)
    # We use 10000 because price diff of 0.0001 * 10000 = $1 profit per lot
    'EUR/USD': 100000,
    'EURUSD': 100000,
    'GBP/USD': 100000,
    'GBPUSD': 100000,
    'AUD/USD': 100000,
    'AUDUSD': 100000,
    'NZD/USD': 100000,
    'NZDUSD': 100000,
    'USD/CAD': 100000,
    'USDCAD': 100000,
    'USD/CHF': 100000,
    'USDCHF': 100000,
    'USD/JPY': 1000,    # JPY pairs have different pip value
    'USDJPY': 1000,
    
    # Crypto - Usually 1:1 or varies
    'BTC/USD': 1,       # 1 contract = 1 BTC
    'BTCUSD': 1,
    'BTC': 1,
    'ETH/USD': 1,       # 1 contract = 1 ETH
    'ETHUSD': 1,
    'ETH': 1,
    
    # Indices - Point value per lot
    'NAS100': 1,        # $1 per point per lot
    'US30': 1,
    'SPX500': 1,
    'US500': 1,
    'Stocks': 1,
}

def get_instrument_multiplier(instrument: str) -> float:
    """Get the contract multiplier for an instrument"""
    # Check exact match first
    if instrument in INSTRUMENT_MULTIPLIERS:
        return INSTRUMENT_MULTIPLIERS[instrument]
    
    # Check partial matches (for variations like "GOLD", "gold", etc.)
    instrument_upper = instrument.upper()
    for key, value in INSTRUMENT_MULTIPLIERS.items():
        if key.upper() in instrument_upper or instrument_upper in key.upper():
            return value
    
    # Default to 1 for unknown instruments
    return 1

def calculate_pnl(trade: dict) -> dict:
    if trade.get('exit_price') and trade.get('status') == 'closed':
        # For MT5-synced trades, keep the PnL from MetaApi (already converted to USD)
        if trade.get('mt5_ticket') and trade.get('pnl') is not None:
            # Only calculate percentage if missing
            if trade.get('pnl_percentage') is None:
                entry = trade['entry_price']
                exit_p = trade['exit_price']
                position = trade['position']
                pnl_percentage = ((exit_p - entry) / entry * 100) if position in ['long', 'buy'] else ((entry - exit_p) / entry * 100)
                trade['pnl_percentage'] = round(pnl_percentage, 2)
            return trade
        
        entry = trade['entry_price']
        exit_p = trade['exit_price']
        qty = trade['quantity']  # This is lot size
        position = trade['position']
        commission = trade.get('commission', 0) or 0
        swap = trade.get('swap', 0) or 0
        instrument = trade.get('instrument', '')
        
        # Get the multiplier for this instrument
        multiplier = get_instrument_multiplier(instrument)
        
        # Calculate raw P&L based on position
        if position in ['long', 'buy']:
            raw_pnl = (exit_p - entry) * qty * multiplier
        else:  # short/sell
            raw_pnl = (entry - exit_p) * qty * multiplier
        
        # Subtract trading costs
        pnl = raw_pnl - commission - swap
        
        # Calculate percentage based on entry price
        pnl_percentage = ((exit_p - entry) / entry * 100) if position in ['long', 'buy'] else ((entry - exit_p) / entry * 100)
        
        trade['pnl'] = round(pnl, 2)
        trade['pnl_percentage'] = round(pnl_percentage, 2)
    else:
        trade['pnl'] = None
        trade['pnl_percentage'] = None
    return trade


def normalize_position_value(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    normalized = str(value).strip().lower()
    if normalized in {'long', 'buy'}:
        return 'buy'
    if normalized in {'short', 'sell'}:
        return 'sell'
    return normalized


def sanitize_trade_for_response(trade: dict) -> dict:
    sanitized = dict(trade)
    sanitized['position'] = normalize_position_value(sanitized.get('position')) or 'buy'

    status_value = str(sanitized.get('status') or '').strip().lower()
    if status_value not in {'open', 'closed'}:
        status_value = 'closed' if sanitized.get('exit_price') is not None else 'open'
    sanitized['status'] = status_value

    if sanitized.get('notes') is None:
        sanitized['notes'] = ''

    if sanitized.get('created_at') is None:
        sanitized['created_at'] = datetime.now(timezone.utc).isoformat()

    if sanitized.get('entry_date') is None:
        sanitized['entry_date'] = sanitized['created_at'][:19]

    if sanitized.get('commission') is None:
        sanitized['commission'] = 0

    if sanitized.get('swap') is None:
        sanitized['swap'] = 0

    return calculate_pnl(sanitized)


async def get_users_by_ids(user_ids: List[str], projection: Optional[Dict[str, int]] = None) -> Dict[str, dict]:
    unique_ids = [user_id for user_id in set(user_ids) if user_id]
    if not unique_ids:
        return {}

    user_projection = {"_id": 0, "id": 1}
    if projection:
        user_projection.update(projection)

    users = await db.users.find({"id": {"$in": unique_ids}}, user_projection).to_list(len(unique_ids))
    return {user["id"]: user for user in users}


async def get_trade_summaries_by_user(user_ids: List[str]) -> Dict[str, Dict[str, Any]]:
    unique_ids = [user_id for user_id in set(user_ids) if user_id]
    summaries = {
        user_id: {"trade_count": 0, "total_pnl": 0.0}
        for user_id in unique_ids
    }
    if not unique_ids:
        return summaries

    trades = await db.trades.find(
        {"user_id": {"$in": unique_ids}},
        {
            "_id": 0,
            "user_id": 1,
            "status": 1,
            "entry_price": 1,
            "exit_price": 1,
            "quantity": 1,
            "position": 1,
            "commission": 1,
            "swap": 1,
            "instrument": 1,
        },
    ).to_list(100000)

    for trade in trades:
        user_id = trade.get("user_id")
        if not user_id:
            continue
        summary = summaries.setdefault(user_id, {"trade_count": 0, "total_pnl": 0.0})
        summary["trade_count"] += 1
        if trade.get("status") == "closed":
            summary["total_pnl"] += calculate_pnl(trade).get("pnl", 0) or 0

    for summary in summaries.values():
        summary["total_pnl"] = round(summary["total_pnl"], 2)

    return summaries

# ============ AUTH ROUTES ============

@api_router.post("/auth/register")
async def register(user_data: UserRegister):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    trial_end = datetime.now(timezone.utc) + timedelta(days=TRIAL_DURATION_DAYS)
    
    # Create user with trial already included
    user_doc = {
        "id": user_id,
        "email": user_data.email,
        "name": user_data.name,
        "password": hash_password(user_data.password),
        "xp_balance": 0,
        "xp_updated_at": now,
        "subscription_status": "trial",
        "subscription_plan": None,
        "subscription_start_date": now,
        "subscription_end_date": trial_end.isoformat(),
        "role": "user",
        "status": "active",
        "created_at": now
    }
    
    await db.users.insert_one(user_doc)
    
    # Track referral if code provided (reward only on payment, not signup)
    if user_data.referral_code:
        try:
            await referral_service.track_signup_from_referral(
                new_user_id=user_id,
                referral_code=user_data.referral_code.upper()
            )
            logging.info(f"Referral tracked for new user: {user_id}")
        except Exception as e:
            logging.error(f"Failed to track referral: {str(e)}")
            # Don't fail registration if referral tracking fails
    
    token = create_access_token(user_id, user_data.email)
    
    # Include subscription status in register response too
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user_id,
            "email": user_data.email,
            "name": user_data.name,
            "created_at": now
        },
        "subscription": {
            "user_id": user_id,
            "subscription_status": "trial",
            "subscription_plan": None,
            "subscription_start_date": now,
            "subscription_end_date": trial_end.isoformat(),
            "is_active": True  # Trial is always active initially
        }
    }

@api_router.post("/auth/login")
async def login(user_data: UserLogin):
    user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if not user or not verify_password(user_data.password, user['password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_access_token(user['id'], user['email'])
    
    # Include subscription status in login response to avoid frontend delay
    is_active = await check_subscription_status(user)
    
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user['id'],
            "email": user['email'],
            "name": user['name'],
            "created_at": user['created_at']
        },
        "subscription": {
            "user_id": user['id'],
            "subscription_status": user.get('subscription_status', 'expired'),
            "subscription_plan": user.get('subscription_plan'),
            "subscription_start_date": user.get('subscription_start_date'),
            "subscription_end_date": user.get('subscription_end_date'),
            "is_active": is_active
        }
    }

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(**current_user)

# ============ FORGOT PASSWORD ============

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """Send password reset email"""
    user = await db.users.find_one({"email": request.email}, {"_id": 0})
    
    # Always return success to prevent email enumeration
    if not user:
        return {"message": "If this email exists, a reset link has been sent"}
    
    # Generate reset token
    reset_token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(hours=1)
    
    # Store reset token
    await db.password_resets.insert_one({
        "user_id": user['id'],
        "email": request.email,
        "token": reset_token,
        "expires_at": expires_at.isoformat(),
        "used": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Send email if Resend is configured
    if RESEND_API_KEY:
        reset_link = f"{FRONTEND_URL}/reset-password?token={reset_token}"
        
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <h2 style="color: #10B981;">TradeLedger Password Reset</h2>
            <p>Hi {user['name']},</p>
            <p>You requested to reset your password. Click the button below to create a new password:</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_link}" style="background-color: #10B981; color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; font-weight: bold;">
                    Reset Password
                </a>
            </div>
            <p>Or copy this link: <a href="{reset_link}">{reset_link}</a></p>
            <p style="color: #666; font-size: 12px;">This link expires in 1 hour. If you didn't request this, please ignore this email.</p>
        </div>
        """
        
        try:
            params = {
                "from": SENDER_EMAIL,
                "to": [request.email],
                "subject": "Reset Your TradeLedger Password",
                "html": html_content
            }
            await asyncio.to_thread(resend.Emails.send, params)
        except Exception as e:
            logging.error(f"Failed to send reset email: {str(e)}")
    
    return {"message": "If this email exists, a reset link has been sent", "token": reset_token if not RESEND_API_KEY else None}

@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    """Reset password using token"""
    reset_doc = await db.password_resets.find_one({
        "token": request.token,
        "used": False
    }, {"_id": 0})
    
    if not reset_doc:
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")
    
    # Check expiration
    expires_at = datetime.fromisoformat(reset_doc['expires_at'].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Reset token has expired")
    
    # Update password
    new_hashed = hash_password(request.new_password)
    await db.users.update_one(
        {"id": reset_doc['user_id']},
        {"$set": {"password": new_hashed}}
    )
    
    # Mark token as used
    await db.password_resets.update_one(
        {"token": request.token},
        {"$set": {"used": True}}
    )
    
    return {"message": "Password reset successfully"}

@api_router.get("/auth/verify-reset-token")
async def verify_reset_token(token: str):
    """Verify if reset token is valid"""
    reset_doc = await db.password_resets.find_one({
        "token": token,
        "used": False
    }, {"_id": 0})
    
    if not reset_doc:
        raise HTTPException(status_code=400, detail="Invalid reset token")
    
    expires_at = datetime.fromisoformat(reset_doc['expires_at'].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Reset token has expired")
    
    return {"valid": True, "email": reset_doc['email']}

# ============ REFERRAL ENDPOINTS ============
create_referral_endpoints(api_router, referral_service, get_current_user, FRONTEND_URL)

# ============ SUBSCRIPTION ENDPOINTS ============

@api_router.get("/subscriptions/my-subscription")
async def get_my_subscription(current_user: dict = Depends(get_current_user)):
    """Get current user subscription status"""
    is_active = await check_subscription_status(current_user)
    return {
        "user_id": current_user['id'],
        "subscription_status": current_user.get('subscription_status', 'expired'),
        "subscription_plan": current_user.get('subscription_plan'),
        "subscription_start_date": current_user.get('subscription_start_date'),
        "subscription_end_date": current_user.get('subscription_end_date'),
        "is_active": is_active
    }

# ============ ADMIN ROUTES ============

@admin_router.post("/login", response_model=AdminTokenResponse)
async def admin_login(user_data: UserLogin):
    """Admin login endpoint"""
    if user_data.email != ADMIN_EMAIL or user_data.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid admin credentials")
    
    token = create_access_token("admin", ADMIN_EMAIL, is_admin=True)
    return AdminTokenResponse(access_token=token)

@admin_router.get("/stats")
async def get_admin_stats(admin: dict = Depends(get_admin_user)):
    """Get overall platform statistics"""
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    total_users, total_trades, total_mt5_accounts, all_trades, recent_users, recent_trades = await asyncio.gather(
        db.users.count_documents({}),
        db.trades.count_documents({}),
        db.mt5_accounts.count_documents({}),
        db.trades.find({"status": "closed"}, {"_id": 0}).to_list(100000),
        db.users.count_documents({"created_at": {"$gte": week_ago}}),
        db.trades.count_documents({"created_at": {"$gte": week_ago}}),
    )
    total_pnl = sum(calculate_pnl(t).get('pnl', 0) or 0 for t in all_trades)

    return {
        "total_users": total_users,
        "total_trades": total_trades,
        "total_mt5_accounts": total_mt5_accounts,
        "total_pnl": round(total_pnl, 2),
        "recent_users_7d": recent_users,
        "recent_trades_7d": recent_trades
    }

@admin_router.get("/users")
async def get_all_users(
    skip: int = 0,
    limit: int = 50,
    admin: dict = Depends(get_admin_user)
):
    """Get all registered users"""
    users = await db.users.find({}, {"_id": 0, "password": 0}).skip(skip).limit(limit).sort("created_at", -1).to_list(limit)

    trade_summaries = await get_trade_summaries_by_user([user['id'] for user in users])
    enriched_users = [
        {
            **user,
            "trade_count": trade_summaries.get(user['id'], {}).get("trade_count", 0),
            "total_pnl": trade_summaries.get(user['id'], {}).get("total_pnl", 0.0),
        }
        for user in users
    ]

    total = await db.users.count_documents({})
    return {"users": enriched_users, "total": total}

@admin_router.get("/users/{user_id}")
async def get_user_details(user_id: str, admin: dict = Depends(get_admin_user)):
    """Get detailed info about a specific user"""
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get user's trades
    trades = await db.trades.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    trades = [calculate_pnl(t) for t in trades]
    
    # Get user's MT5 accounts
    mt5_accounts = await db.mt5_accounts.find({"user_id": user_id}, {"_id": 0, "password": 0}).to_list(100)
    
    # Calculate stats
    closed_trades = [t for t in trades if t['status'] == 'closed']
    total_pnl = sum(t.get('pnl', 0) or 0 for t in closed_trades)
    winning = len([t for t in closed_trades if (t.get('pnl') or 0) > 0])
    losing = len([t for t in closed_trades if (t.get('pnl') or 0) < 0])
    
    return {
        "user": user,
        "trades": trades[:50],  # Last 50 trades
        "mt5_accounts": mt5_accounts,
        "stats": {
            "total_trades": len(trades),
            "closed_trades": len(closed_trades),
            "open_trades": len(trades) - len(closed_trades),
            "total_pnl": round(total_pnl, 2),
            "winning_trades": winning,
            "losing_trades": losing,
            "win_rate": round(winning / len(closed_trades) * 100, 2) if closed_trades else 0
        }
    }

@admin_router.get("/trades")
async def get_all_trades(
    skip: int = 0,
    limit: int = 100,
    admin: dict = Depends(get_admin_user)
):
    """Get all trades from all users"""
    trades = await db.trades.find({}, {"_id": 0}).skip(skip).limit(limit).sort("created_at", -1).to_list(limit)
    trades = [calculate_pnl(t) for t in trades]

    user_map = await get_users_by_ids([trade.get('user_id') for trade in trades], {"name": 1, "email": 1})
    enriched_trades = [
        {
            **trade,
            "user_name": user_map.get(trade.get('user_id'), {}).get('name', 'Unknown'),
            "user_email": user_map.get(trade.get('user_id'), {}).get('email', 'Unknown'),
        }
        for trade in trades
    ]

    total = await db.trades.count_documents({})
    return {"trades": enriched_trades, "total": total}

@admin_router.get("/activity")
async def get_recent_activity(
    limit: int = 50,
    admin: dict = Depends(get_admin_user)
):
    """Get recent platform activity"""
    # Recent user registrations
    recent_users = await db.users.find({}, {"_id": 0, "password": 0}).sort("created_at", -1).limit(10).to_list(10)
    
    # Recent trades
    recent_trades = await db.trades.find({}, {"_id": 0}).sort("created_at", -1).limit(20).to_list(20)
    recent_trades = [calculate_pnl(t) for t in recent_trades]

    user_map = await get_users_by_ids([trade.get('user_id') for trade in recent_trades], {"name": 1})
    for trade in recent_trades:
        trade['user_name'] = user_map.get(trade.get('user_id'), {}).get('name', 'Unknown')

    # Recent password resets
    recent_resets = await db.password_resets.find({}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    
    return {
        "recent_users": recent_users,
        "recent_trades": recent_trades,
        "recent_password_resets": recent_resets
    }

@admin_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(get_admin_user)):
    """Delete a user and all their data"""
    # Check if user exists
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Delete user's trades
    await db.trades.delete_many({"user_id": user_id})
    
    # Delete user's MT5 accounts
    await db.mt5_accounts.delete_many({"user_id": user_id})
    
    # Delete user's password resets
    await db.password_resets.delete_many({"user_id": user_id})
    
    # Delete the user
    await db.users.delete_one({"id": user_id})
    
    return {"message": f"User {user['email']} and all associated data deleted successfully"}

@admin_router.get("/export/trades")
async def export_all_trades_csv(admin: dict = Depends(get_admin_user)):
    """Export all trades from all users as CSV"""
    trades = await db.trades.find({}, {"_id": 0}).sort("created_at", -1).to_list(100000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row
    writer.writerow([
        'User ID', 'Instrument', 'Position', 'Status', 'Entry Price', 'Exit Price', 
        'Quantity', 'Entry Date', 'Exit Date', 'P&L', 'P&L %',
        'Stop Loss', 'Take Profit', 'Commission', 'Swap', 'Notes', 'Created At'
    ])
    
    # Data rows
    for trade in trades:
        trade = calculate_pnl(trade)
        writer.writerow([
            trade.get('user_id', ''),
            trade.get('instrument', ''),
            trade.get('position', ''),
            trade.get('status', ''),
            trade.get('entry_price', ''),
            trade.get('exit_price', ''),
            trade.get('quantity', ''),
            trade.get('entry_date', ''),
            trade.get('exit_date', ''),
            trade.get('pnl', ''),
            trade.get('pnl_percentage', ''),
            trade.get('stop_loss', ''),
            trade.get('take_profit', ''),
            trade.get('commission', ''),
            trade.get('swap', ''),
            trade.get('notes', ''),
            trade.get('created_at', '')
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=all_trades_export_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


# ============ ADMIN DATABASE ENDPOINTS ============

@admin_router.get("/database/overview")
async def get_database_overview(admin: dict = Depends(get_admin_user)):
    '''Get comprehensive overview of all data in the database'''
    try:
        collections_info = {}
        users_count = await db.users.count_documents({})
        users_sample = await db.users.find({}, {"_id": 0, "password": 0}).limit(5).to_list(5)
        collections_info['users'] = {"count": users_count, "sample": users_sample}
        trades_count = await db.trades.count_documents({})
        trades_sample = await db.trades.find({}, {"_id": 0}).limit(5).to_list(5)
        collections_info['trades'] = {"count": trades_count, "sample": trades_sample}
        mt5_count = await db.mt5_accounts.count_documents({})
        mt5_sample = await db.mt5_accounts.find({}, {"_id": 0, "password": 0}).limit(5).to_list(5)
        collections_info['mt5_accounts'] = {"count": mt5_count, "sample": mt5_sample}
        return {"database_name": db.name, "collections": collections_info, "total_users": users_count, "total_trades": trades_count, "total_mt5_accounts": mt5_count}
    except Exception as e:
        logging.error(f"Database overview error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch database overview: {str(e)}")

@admin_router.get("/database/users/all")
async def get_all_users_data(admin: dict = Depends(get_admin_user), limit: int = Query(100, ge=1, le=1000), skip: int = Query(0, ge=0)):
    '''Get all users with statistics'''
    try:
        users = await db.users.find({}, {"_id": 0, "password": 0}).skip(skip).limit(limit).to_list(limit)
        for user in users:
            total_trades = await db.trades.count_documents({"user_id": user['id']})
            user['statistics'] = {"total_trades": total_trades}
        total_users = await db.users.count_documents({})
        return {"total_users": total_users, "returned": len(users), "users": users}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch users: {str(e)}")

@admin_router.get("/database/trades/all")
async def get_all_trades_data(admin: dict = Depends(get_admin_user), limit: int = Query(100, ge=1, le=1000), skip: int = Query(0, ge=0)):
    '''Get all trades'''
    try:
        trades = await db.trades.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
        total_trades = await db.trades.count_documents({})
        return {"total_trades": total_trades, "returned": len(trades), "trades": trades}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch trades: {str(e)}")

# ============ ADMIN REFERRAL ENDPOINTS ============
create_admin_referral_endpoints(admin_router, referral_service, db, get_admin_user)

# ============ HEALTH & DEBUG ENDPOINTS ============

@api_router.get("/health/ai-config")
async def check_ai_config():
    """Check AI service configuration status"""
    return {
        "xai_configured": bool(XAI_API_KEY),
        "xai_key_length": len(XAI_API_KEY) if XAI_API_KEY else 0,
        "xai_key_prefix": XAI_API_KEY[:10] + "..." if XAI_API_KEY and len(XAI_API_KEY) > 10 else "NOT_SET",
        "groq_configured": bool(GROQ_API_KEY),
        "groq_key_length": len(GROQ_API_KEY) if GROQ_API_KEY else 0,
        "xai_grok_enabled": xai_grok_enabled,
        "groq_client_available": groq_client is not None
    }

# ============ MT5 ACCOUNT ROUTES ============

@api_router.post("/mt5/metaapi-token")
async def save_metaapi_token(current_user: dict = Depends(get_current_user), body: dict = {}):
    """Save MetaApi API token for the user"""
    token = body.get('token', '').strip()
    if not token:
        raise HTTPException(status_code=400, detail="Token is required")
    await db.user_settings.update_one(
        {"user_id": current_user['id']},
        {"$set": {"user_id": current_user['id'], "metaapi_token": token, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"message": "MetaApi token saved successfully"}

@api_router.get("/mt5/metaapi-token")
async def get_metaapi_token(current_user: dict = Depends(get_current_user)):
    """Check if user has a MetaApi token saved"""
    settings = await db.user_settings.find_one({"user_id": current_user['id']}, {"_id": 0})
    has_token = bool(settings and settings.get('metaapi_token'))
    return {"has_token": has_token}

@api_router.delete("/mt5/metaapi-token")
async def delete_metaapi_token(current_user: dict = Depends(get_current_user)):
    """Remove saved MetaApi token"""
    await db.user_settings.update_one(
        {"user_id": current_user['id']},
        {"$unset": {"metaapi_token": ""}}
    )
    return {"message": "MetaApi token removed"}

@api_router.post("/mt5/accounts", response_model=MT5AccountResponse)
async def add_mt5_account(account_data: MT5AccountCreate, current_user: dict = Depends(get_current_user)):
    account_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    account_doc = {
        "id": account_id,
        "user_id": current_user['id'],
        "name": account_data.name,
        "login": account_data.login,
        "password": account_data.password,
        "server": account_data.server,
        "platform": account_data.platform,
        "metaapi_account_id": account_data.metaapi_account_id or None,
        "currency": account_data.currency or "USD",
        "is_connected": False,
        "last_sync": None,
        "balance": None,
        "equity": None,
        "created_at": now
    }
    
    await db.mt5_accounts.insert_one(account_doc)
    
    del account_doc['password']
    if '_id' in account_doc:
        del account_doc['_id']
    
    return MT5AccountResponse(**account_doc)

@api_router.get("/mt5/accounts", response_model=List[MT5AccountResponse])
async def get_mt5_accounts(current_user: dict = Depends(get_current_user)):
    accounts = await db.mt5_accounts.find(
        {"user_id": current_user['id']},
        {"_id": 0, "password": 0}
    ).to_list(100)
    return [MT5AccountResponse(**acc) for acc in accounts]

@api_router.delete("/mt5/accounts/{account_id}")
async def delete_mt5_account(account_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.mt5_accounts.delete_one({"id": account_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Account not found")
    return {"message": "Account deleted successfully"}

@api_router.put("/mt5/accounts/{account_id}", response_model=MT5AccountResponse)
async def update_mt5_account(account_id: str, account_data: MT5AccountUpdate, current_user: dict = Depends(get_current_user)):
    """Update MT5 account (only currency and metaapi_account_id can be changed)"""
    # Verify account exists and belongs to user
    existing = await db.mt5_accounts.find_one({"id": account_id, "user_id": current_user['id']})
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Build update fields
    update_fields = {}
    if account_data.currency is not None:
        update_fields["currency"] = account_data.currency
    if account_data.metaapi_account_id is not None:
        update_fields["metaapi_account_id"] = account_data.metaapi_account_id
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    # Update the account
    await db.mt5_accounts.update_one(
        {"id": account_id, "user_id": current_user['id']},
        {"$set": update_fields}
    )
    
    # Fetch and return updated account
    updated_account = await db.mt5_accounts.find_one({"id": account_id}, {"_id": 0, "password": 0})
    return MT5AccountResponse(**updated_account)

@api_router.post("/mt5/accounts/{account_id}/sync")
async def sync_mt5_account(account_id: str, current_user: dict = Depends(get_current_user)):
    """Sync trades from MT5 account via MetaApi REST API"""
    account = await db.mt5_accounts.find_one(
        {"id": account_id, "user_id": current_user['id']},
        {"_id": 0}
    )
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    now = datetime.now(timezone.utc).isoformat()
    metaapi_id = account.get('metaapi_account_id')
    
    # Get MetaApi token: user's saved token first, then env var fallback
    user_settings = await db.user_settings.find_one({"user_id": current_user['id']}, {"_id": 0})
    metaapi_token = (user_settings or {}).get('metaapi_token') or METAAPI_TOKEN
    
    if not metaapi_token:
        raise HTTPException(status_code=503, detail="MetaApi token not configured. Go to Accounts page and add your MetaApi API Token.")
    
    if not metaapi_id:
        raise HTTPException(status_code=400, detail="MetaApi Account ID is not set for this account. Edit the account and add your MetaApi Account ID.")
    
    try:
        headers = {"auth-token": metaapi_token}
        
        async with httpx.AsyncClient(timeout=60, verify=False) as http:
            # Step 1: Get account info from MetaApi provisioning API
            prov_url = f"https://mt-provisioning-api-v1.agiliumtrade.agiliumtrade.ai/users/current/accounts/{metaapi_id}"
            acc_resp = await http.get(prov_url, headers=headers)
            if acc_resp.status_code == 404:
                raise HTTPException(status_code=400, detail="MetaApi Account ID not found. Check the ID in your MetaApi dashboard.")
            acc_resp.raise_for_status()
            acc_info = acc_resp.json()
            
            region = acc_info.get('region', 'vint-hill')
            acc_state = acc_info.get('state', 'UNDEPLOYED')
            
            # Step 2: Deploy account if needed
            if acc_state == 'UNDEPLOYED':
                try:
                    deploy_resp = await http.post(f"{prov_url}/deploy", headers=headers)
                    deploy_resp.raise_for_status()
                    # Wait briefly for deployment
                    for _ in range(10):
                        await asyncio.sleep(3)
                        check = await http.get(prov_url, headers=headers)
                        check_data = check.json()
                        if check_data.get('connectionStatus') == 'CONNECTED':
                            break
                except Exception as deploy_err:
                    logging.warning(f"MetaApi deploy failed (may need manual deploy): {deploy_err}")
                    raise HTTPException(
                        status_code=400,
                        detail="Account is not deployed. Please log in to your MetaApi dashboard at https://app.metaapi.cloud, deploy the account manually, then sync again. Your API token may not have deploy permissions — use a token with full access or upgrade your MetaApi plan."
                    )
            
            # Step 3: Fetch history deals from MetaApi client API (paginated, max 1000 per page)
            start_time = (datetime.now(timezone.utc) - timedelta(days=730)).strftime('%Y-%m-%dT%H:%M:%S.000Z')
            end_time = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')
            
            client_base = f"https://mt-client-api-v1.{region}.agiliumtrade.ai"
            deals_url = f"{client_base}/users/current/accounts/{metaapi_id}/history-deals/time/{start_time}/{end_time}"
            
            deals = []
            offset = 0
            limit = 1000
            while True:
                resp = await http.get(deals_url, headers=headers, params={"offset": offset, "limit": limit})
                resp.raise_for_status()
                page = resp.json()
                if not page:
                    break
                deals.extend(page)
                if len(page) < limit:
                    break
                offset += limit
        
        if not deals:
            await db.mt5_accounts.update_one(
                {"id": account_id},
                {"$set": {"last_sync": now, "is_connected": True}}
            )
            return {"message": "No deals found in MetaApi history.", "new_trades": 0, "total_deals": 0}
        
        # Step 4: Group deals by positionId to reconstruct trades
        # Each trade has an "in" deal (entry) and an "out" deal (exit/close)
        # Skip balance operations and deals without a symbol
        positions = {}
        for deal in deals:
            pos_id = deal.get('positionId')
            symbol = deal.get('symbol', '')
            if not pos_id or not symbol:
                continue
            # Skip non-trade deal types (balance, credit, etc.)
            deal_type = deal.get('type', '')
            if 'BALANCE' in deal_type or 'CREDIT' in deal_type or 'CHARGE' in deal_type:
                continue
            entry_type = deal.get('entryType', '')
            if pos_id not in positions:
                positions[pos_id] = {'entries': [], 'exits': []}
            if entry_type == 'DEAL_ENTRY_IN':
                positions[pos_id]['entries'].append(deal)
            elif entry_type == 'DEAL_ENTRY_OUT':
                positions[pos_id]['exits'].append(deal)
        
        # Get account base currency for conversion
        base_currency = acc_info.get('baseCurrency', 'USD')
        
        # Step 5: Build trade documents
        trades_to_import = []
        for pos_id, pos_deals in positions.items():
            if not pos_deals['entries']:
                continue
            
            entry = pos_deals['entries'][0]
            exit_deal = pos_deals['exits'][0] if pos_deals['exits'] else None
            
            raw_symbol = entry.get('symbol', '')
            normalized_symbol = normalize_symbol(raw_symbol)
            
            # Direction comes from the "in" deal's type: BUY = long, SELL = short
            in_type = entry.get('type', '')
            direction = 'buy' if 'BUY' in in_type else 'sell'
            
            # Commission and swap from BOTH in and out deals
            all_deals_for_pos = pos_deals['entries'] + pos_deals['exits']
            total_commission = sum(d.get('commission', 0) or 0 for d in all_deals_for_pos)
            total_swap = sum(d.get('swap', 0) or 0 for d in all_deals_for_pos)
            
            # PnL: ONLY from the "out" (close) deal's profit field
            pnl = None
            # Use the currency setting from the account (INR or USD)
            account_currency = account.get('currency', 'USD')
            if exit_deal:
                raw_profit = exit_deal.get('profit', 0) or 0
                raw_pnl = raw_profit + total_commission + total_swap
                
                if account_currency == 'USD' and base_currency != 'USD':
                    # Convert from account currency to USD
                    exchange_rate = exit_deal.get('accountCurrencyExchangeRate', None)
                    if exchange_rate and exchange_rate > 0:
                        pnl = round(raw_pnl / exchange_rate, 2)
                    else:
                        pnl = round(raw_pnl, 2)
                else:
                    # Keep raw value (INR or already USD)
                    pnl = round(raw_pnl, 2)
            
            entry_time = entry.get('time', '')
            exit_time = exit_deal.get('time', '') if exit_deal else None
            
            # Entry price from "in" deal, exit price from "out" deal
            entry_price = entry.get('price', 0)
            exit_price = exit_deal.get('price') if exit_deal else None
            
            trade = {
                "id": str(uuid.uuid4()),
                "user_id": current_user['id'],
                "instrument": normalized_symbol,
                "position": direction,
                "entry_price": entry_price,
                "exit_price": exit_price,
                "quantity": entry.get('volume', 0),
                "entry_date": entry_time,
                "exit_date": exit_time,
                "stop_loss": None,
                "take_profit": None,
                "status": "closed" if exit_deal else "open",
                "pnl": pnl,
                "commission": round(total_commission, 2),
                "swap": round(total_swap, 2),
                "notes": f"Auto-synced from MT5 (position: {pos_id})",
                "mt5_ticket": str(pos_id),
                "mt5_account_id": account_id,
                "currency": account_currency,
                "created_at": now,
                "updated_at": now,
            }
            trades_to_import.append(trade)
        
        # Step 6: Deduplicate — skip trades whose mt5_ticket already exists
        existing = await db.trades.find(
            {"user_id": current_user['id'], "mt5_ticket": {"$ne": None}},
            {"mt5_ticket": 1}
        ).to_list(None)
        existing_tickets = {t['mt5_ticket'] for t in existing}
        
        new_trades = [t for t in trades_to_import if t['mt5_ticket'] not in existing_tickets]
        
        # Step 7: Insert new trades
        if new_trades:
            await db.trades.insert_many(new_trades)
        
        # Step 8: Update account status
        update_fields = {
            "is_connected": True,
            "last_sync": now,
        }
        # Try to get balance/equity from account info
        if 'balance' in acc_info:
            update_fields['balance'] = acc_info.get('balance')
        if 'equity' in acc_info:
            update_fields['equity'] = acc_info.get('equity')
        
        await db.mt5_accounts.update_one(
            {"id": account_id},
            {"$set": update_fields}
        )
        
        return {
            "message": f"Sync completed! {len(new_trades)} new trades imported.",
            "total_deals": len(deals),
            "total_positions": len(positions),
            "new_trades": len(new_trades),
            "skipped_duplicates": len(trades_to_import) - len(new_trades),
        }
        
    except HTTPException:
        raise
    except httpx.HTTPStatusError as e:
        logging.error(f"MetaApi HTTP error: {e.response.status_code} - {e.response.text}")
        error_detail = "MetaApi API error"
        try:
            err_body = e.response.json()
            error_detail = err_body.get('message', str(e))
        except Exception:
            error_detail = e.response.text[:200]
        raise HTTPException(status_code=502, detail=f"MetaApi error: {error_detail}")
    except Exception as e:
        logging.error(f"MT5 sync error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")

# ============ TRADE ROUTES ============

@api_router.post("/trades", response_model=TradeResponse)
async def create_trade(trade_data: TradeCreate, current_user: dict = Depends(get_current_user)):
    trade_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    trade_doc = {
        "id": trade_id,
        "user_id": current_user['id'],
        **trade_data.model_dump(),
        "created_at": now
    }
    trade_doc['position'] = normalize_position_value(trade_doc.get('position'))
    
    trade_doc = sanitize_trade_for_response(trade_doc)
    await db.trades.insert_one(trade_doc)
    
    if '_id' in trade_doc:
        del trade_doc['_id']
    return TradeResponse(**trade_doc)

@api_router.get("/trades", response_model=List[TradeResponse])
async def get_trades(
    status: Optional[str] = None,
    instrument: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user['id']}
    if status:
        query["status"] = status
    if instrument:
        query["instrument"] = instrument
    if start_date:
        query["entry_date"] = {"$gte": start_date}
    if end_date:
        query.setdefault("entry_date", {})["$lte"] = end_date
    
    trades = await db.trades.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    response_trades = []
    invalid_rows = 0
    for trade in trades:
        try:
            response_trades.append(TradeResponse(**sanitize_trade_for_response(trade)))
        except Exception as e:
            invalid_rows += 1
            logging.warning(f"Skipping invalid trade row id={trade.get('id')}: {str(e)}")

    if invalid_rows:
        logging.warning(f"Skipped {invalid_rows} invalid trade rows for user {current_user.get('id')}")

    return response_trades

@api_router.get("/trades/{trade_id}", response_model=TradeResponse)
async def get_trade(trade_id: str, current_user: dict = Depends(get_current_user)):
    trade = await db.trades.find_one({"id": trade_id, "user_id": current_user['id']}, {"_id": 0})
    if not trade:
        raise HTTPException(status_code=404, detail="Trade not found")
    return TradeResponse(**sanitize_trade_for_response(trade))

@api_router.put("/trades/{trade_id}", response_model=TradeResponse)
async def update_trade(trade_id: str, trade_data: TradeUpdate, current_user: dict = Depends(get_current_user)):
    existing = await db.trades.find_one({"id": trade_id, "user_id": current_user['id']}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Trade not found")
    
    update_dict = {k: v for k, v in trade_data.model_dump().items() if v is not None}
    if 'position' in update_dict:
        update_dict['position'] = normalize_position_value(update_dict.get('position'))
    if update_dict:
        await db.trades.update_one({"id": trade_id}, {"$set": update_dict})
    
    updated = await db.trades.find_one({"id": trade_id}, {"_id": 0})
    return TradeResponse(**sanitize_trade_for_response(updated))

@api_router.delete("/trades/{trade_id}")
async def delete_trade(trade_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.trades.delete_one({"id": trade_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Trade not found")
    return {"message": "Trade deleted successfully"}

@api_router.get("/leaderboard")
async def get_leaderboard(limit: int = 10, current_user: dict = Depends(get_current_user)):
    """
    Get leaderboard of top traders based on performance score.
    Score = (win_rate * 0.4) + (avg_rr * 20 * 0.3) + (normalized_pnl * 0.3)
    """
    # Get all users
    all_users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(None)
    
    leaderboard_data = []
    
    for user in all_users:
        user_id = user.get('id')
        if not user_id:
            continue
            
        # Get all closed trades for this user
        closed_trades = await db.trades.find({
            "user_id": user_id,
            "status": "closed"
        }, {"_id": 0}).to_list(None)
        
        if not closed_trades:
            continue
            
        # Calculate stats
        total_trades = len(closed_trades)
        wins = sum(1 for t in closed_trades if calculate_pnl(t).get('pnl', 0) > 0)
        losses = sum(1 for t in closed_trades if calculate_pnl(t).get('pnl', 0) < 0)
        win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
        
        # Calculate total PnL
        total_pnl = sum(calculate_pnl(t).get('pnl', 0) or 0 for t in closed_trades)
        
        # Calculate average risk/reward ratio
        rr_values = [t.get('risk_reward', 0) or 0 for t in closed_trades if t.get('risk_reward')]
        avg_rr = sum(rr_values) / len(rr_values) if rr_values else 0
        
        # Normalize PnL (scale to 0-100 range for scoring)
        # We'll use a sigmoid-like function to normalize
        normalized_pnl = min(100, max(0, (total_pnl / 1000) * 50 + 50))
        
        # Calculate performance score
        # win_rate: 0-100, weight 40%
        # avg_rr: 0-5 (scaled to 0-100), weight 30%
        # normalized_pnl: 0-100, weight 30%
        score = (win_rate * 0.4) + (avg_rr * 20 * 0.3) + (normalized_pnl * 0.3)
        
        leaderboard_data.append({
            "user_id": user_id,
            "name": user.get('name', 'Unknown'),
            "score": round(score, 2),
            "total_trades": total_trades,
            "wins": wins,
            "losses": losses,
            "win_rate": round(win_rate, 2),
            "total_pnl": round(total_pnl, 2),
            "avg_rr": round(avg_rr, 2)
        })
    
    # Sort by score descending
    leaderboard_data.sort(key=lambda x: x['score'], reverse=True)
    
    # Add rank
    for idx, entry in enumerate(leaderboard_data[:limit], 1):
        entry['rank'] = idx
    
    return leaderboard_data[:limit]

@api_router.delete("/trades/{trade_id}")
async def delete_trade(trade_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.trades.delete_one({"id": trade_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Trade not found")
    return {"message": "Trade deleted successfully"}

@api_router.delete("/trades")
async def delete_all_trades(current_user: dict = Depends(get_current_user)):
    """Delete all trades for the current user"""
    result = await db.trades.delete_many({"user_id": current_user['id']})
    return {
        "message": f"Successfully deleted {result.deleted_count} trades",
        "deleted_count": result.deleted_count
    }

# ============ ANALYTICS ROUTES ============

@api_router.get("/analytics/summary")
async def get_analytics_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user['id'], "status": "closed"}
    if start_date:
        query["exit_date"] = {"$gte": start_date}
    if end_date:
        query.setdefault("exit_date", {})["$lte"] = end_date
    
    trades = await db.trades.find(query, {"_id": 0}).to_list(10000)
    
    total_trades = len(trades)
    winning_trades = 0
    losing_trades = 0
    total_pnl = 0
    total_wins = 0
    total_losses = 0
    
    for trade in trades:
        trade = calculate_pnl(trade)
        if trade.get('pnl') is not None:
            total_pnl += trade['pnl']
            if trade['pnl'] > 0:
                winning_trades += 1
                total_wins += trade['pnl']
            elif trade['pnl'] < 0:
                losing_trades += 1
                total_losses += abs(trade['pnl'])
    
    win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
    avg_win = total_wins / winning_trades if winning_trades > 0 else 0
    avg_loss = total_losses / losing_trades if losing_trades > 0 else 0
    avg_win_loss_ratio = avg_win / avg_loss if avg_loss > 0 else 0
    
    open_trades = await db.trades.count_documents({"user_id": current_user['id'], "status": "open"})
    
    daily_stats = {}
    for trade in trades:
        if trade.get('exit_date'):
            day = trade['exit_date'][:10]
            if day not in daily_stats:
                daily_stats[day] = {"wins": 0, "losses": 0, "pnl": 0}
            trade = calculate_pnl(trade)
            if trade.get('pnl'):
                daily_stats[day]['pnl'] += trade['pnl']
                if trade['pnl'] > 0:
                    daily_stats[day]['wins'] += 1
                else:
                    daily_stats[day]['losses'] += 1
    
    winning_days = sum(1 for d in daily_stats.values() if d['pnl'] > 0)
    total_days = len(daily_stats)
    daily_win_rate = (winning_days / total_days * 100) if total_days > 0 else 0
    
    day_wins_total = sum(d['pnl'] for d in daily_stats.values() if d['pnl'] > 0)
    day_losses_total = abs(sum(d['pnl'] for d in daily_stats.values() if d['pnl'] < 0))
    day_win_loss_ratio = day_wins_total / day_losses_total if day_losses_total > 0 else 0
    
    sorted_days = sorted(daily_stats.keys())
    current_win_streak = 0
    max_win_streak = 0
    current_trade_streak = 0
    max_trade_streak = 0
    
    for day in sorted_days:
        if daily_stats[day]['pnl'] > 0:
            current_win_streak += 1
            max_win_streak = max(max_win_streak, current_win_streak)
        else:
            current_win_streak = 0
    
    all_trades_sorted = sorted([t for t in trades if t.get('exit_date')], key=lambda x: x['exit_date'])
    for trade in all_trades_sorted:
        trade = calculate_pnl(trade)
        if trade.get('pnl', 0) > 0:
            current_trade_streak += 1
            max_trade_streak = max(max_trade_streak, current_trade_streak)
        else:
            current_trade_streak = 0
    
    # Detect currency from MT5 account first, then from trades
    trade_currency = 'USD'
    mt5_account = await db.mt5_accounts.find_one({"user_id": current_user['id']}, {"_id": 0, "currency": 1})
    if mt5_account and mt5_account.get('currency'):
        trade_currency = mt5_account['currency']
    else:
        # Fallback to currency from trades
        for trade in trades:
            if trade.get('currency'):
                trade_currency = trade['currency']
                break
    
    return {
        "total_trades": total_trades,
        "open_trades": open_trades,
        "winning_trades": winning_trades,
        "losing_trades": losing_trades,
        "win_rate": round(win_rate, 2),
        "total_pnl": round(total_pnl, 2),
        "avg_win": round(avg_win, 2),
        "avg_loss": round(avg_loss, 2),
        "avg_win_loss_ratio": round(avg_win_loss_ratio, 2),
        "daily_win_rate": round(daily_win_rate, 2),
        "day_win_loss_ratio": round(day_win_loss_ratio, 2),
        "trading_days": total_days,
        "winning_days": winning_days,
        "win_streak_days": max_win_streak,
        "win_streak_trades": max_trade_streak,
        "current_win_streak_days": current_win_streak,
        "current_win_streak_trades": current_trade_streak,
        "currency": trade_currency
    }

@api_router.get("/analytics/by-instrument")
async def get_analytics_by_instrument(current_user: dict = Depends(get_current_user)):
    trades = await db.trades.find({"user_id": current_user['id'], "status": "closed"}, {"_id": 0}).to_list(10000)
    
    instrument_stats = {}
    for trade in trades:
        trade = calculate_pnl(trade)
        instrument = trade['instrument']
        
        if instrument not in instrument_stats:
            instrument_stats[instrument] = {
                "instrument": instrument,
                "total_trades": 0,
                "winning_trades": 0,
                "losing_trades": 0,
                "total_pnl": 0,
                "total_volume": 0
            }
        
        instrument_stats[instrument]['total_trades'] += 1
        instrument_stats[instrument]['total_volume'] += trade.get('quantity', 0)
        if trade.get('pnl') is not None:
            instrument_stats[instrument]['total_pnl'] += trade['pnl']
            if trade['pnl'] > 0:
                instrument_stats[instrument]['winning_trades'] += 1
            else:
                instrument_stats[instrument]['losing_trades'] += 1
    
    result = []
    for inst, stats in instrument_stats.items():
        stats['total_pnl'] = round(stats['total_pnl'], 2)
        stats['win_rate'] = round(stats['winning_trades'] / stats['total_trades'] * 100, 2) if stats['total_trades'] > 0 else 0
        result.append(stats)
    
    return sorted(result, key=lambda x: x['total_pnl'], reverse=True)

@api_router.get("/analytics/monthly")
async def get_monthly_analytics(current_user: dict = Depends(get_current_user)):
    trades = await db.trades.find({"user_id": current_user['id'], "status": "closed"}, {"_id": 0}).to_list(10000)
    
    monthly_stats = {}
    for trade in trades:
        trade = calculate_pnl(trade)
        if trade.get('exit_date'):
            month = trade['exit_date'][:7]
            if month not in monthly_stats:
                monthly_stats[month] = {"month": month, "pnl": 0, "trades": 0, "wins": 0}
            
            monthly_stats[month]['trades'] += 1
            if trade.get('pnl') is not None:
                monthly_stats[month]['pnl'] += trade['pnl']
                if trade['pnl'] > 0:
                    monthly_stats[month]['wins'] += 1
    
    result = sorted(monthly_stats.values(), key=lambda x: x['month'])
    for r in result:
        r['pnl'] = round(r['pnl'], 2)
        r['win_rate'] = round(r['wins'] / r['trades'] * 100, 2) if r['trades'] > 0 else 0
    
    return result

@api_router.get("/analytics/daily")
async def get_daily_analytics(
    year: int = Query(default=None),
    month: int = Query(default=None),
    current_user: dict = Depends(get_current_user)
):
    now = datetime.now(timezone.utc)
    year = year or now.year
    month = month or now.month
    
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    
    trades = await db.trades.find({
        "user_id": current_user['id'],
        "status": "closed",
        "exit_date": {"$gte": start_date, "$lt": end_date}
    }, {"_id": 0}).to_list(10000)
    
    daily_stats = {}
    total_trades = 0
    total_wins = 0
    total_pnl = 0
    
    for trade in trades:
        trade = calculate_pnl(trade)
        if trade.get('exit_date'):
            day = trade['exit_date'][:10]
            if day not in daily_stats:
                daily_stats[day] = {"date": day, "pnl": 0, "trades": 0, "wins": 0, "percentage": 0}
            
            daily_stats[day]['trades'] += 1
            total_trades += 1
            if trade.get('pnl') is not None:
                daily_stats[day]['pnl'] += trade['pnl']
                total_pnl += trade['pnl']
                if trade['pnl'] > 0:
                    daily_stats[day]['wins'] += 1
                    total_wins += 1
    
    result = sorted(daily_stats.values(), key=lambda x: x['date'])
    for r in result:
        r['pnl'] = round(r['pnl'], 2)
    
    return {
        "year": year,
        "month": month,
        "days": result,
        "summary": {
            "total_trades": total_trades,
            "total_wins": total_wins,
            "total_pnl": round(total_pnl, 2),
            "win_rate": round(total_wins / total_trades * 100, 2) if total_trades > 0 else 0
        }
    }

@api_router.get("/analytics/balance-history")
async def get_balance_history(
    period: str = Query(default="1M"),
    current_user: dict = Depends(get_current_user)
):
    now = datetime.now(timezone.utc)
    
    period_map = {
        "1D": timedelta(days=1),
        "1W": timedelta(weeks=1),
        "1M": timedelta(days=30),
        "6M": timedelta(days=180),
        "1Y": timedelta(days=365),
        "All": timedelta(days=3650)
    }
    
    start_date = (now - period_map.get(period, timedelta(days=30))).isoformat()[:10]
    
    trades = await db.trades.find({
        "user_id": current_user['id'],
        "status": "closed",
        "exit_date": {"$gte": start_date}
    }, {"_id": 0}).sort("exit_date", 1).to_list(10000)
    
    balance_history = []
    cumulative_pnl = 0
    
    for trade in trades:
        trade = calculate_pnl(trade)
        if trade.get('pnl') is not None and trade.get('exit_date'):
            cumulative_pnl += trade['pnl']
            balance_history.append({
                "date": trade['exit_date'][:10],
                "balance": round(cumulative_pnl, 2),
                "trade_pnl": round(trade['pnl'], 2)
            })
    
    return balance_history

@api_router.get("/analytics/trade-count")
async def get_trade_count_history(
    period: str = Query(default="1M"),
    current_user: dict = Depends(get_current_user)
):
    now = datetime.now(timezone.utc)
    
    period_map = {
        "1D": timedelta(days=1),
        "1W": timedelta(weeks=1),
        "1M": timedelta(days=30),
        "6M": timedelta(days=180),
        "1Y": timedelta(days=365),
        "All": timedelta(days=3650)
    }
    
    start_date = (now - period_map.get(period, timedelta(days=30))).isoformat()[:10]
    
    trades = await db.trades.find({
        "user_id": current_user['id'],
        "entry_date": {"$gte": start_date}
    }, {"_id": 0}).to_list(10000)
    
    daily_counts = {}
    for trade in trades:
        day = trade['entry_date'][:10]
        daily_counts[day] = daily_counts.get(day, 0) + 1
    
    result = sorted([{"date": k, "count": v} for k, v in daily_counts.items()], key=lambda x: x['date'])
    total_count = sum(d['count'] for d in result)
    
    return {"total": total_count, "data": result}


# ============ MARKET DATA ROUTES ============

@api_router.get("/market/candles", response_model=List[CandleResponse])
async def get_market_candles(
    symbol: str = Query(default="BTCUSDT"),
    interval: str = Query(default="5m"),
    limit: int = Query(default=500, ge=50, le=1000),
    from_ts: Optional[int] = Query(default=None, alias="from", ge=0),
    to_ts: Optional[int] = Query(default=None, alias="to", ge=0),
):
    """Fetch OHLCV candles from Twelve Data with Yahoo/Binance fallbacks."""
    interval_map = {
        "1m": "1min",
        "3m": "3min",
        "5m": "5min",
        "15m": "15min",
        "30m": "30min",
        "1h": "1h",
        "2h": "2h",
        "4h": "4h",
        "1d": "1day",
    }
    interval_seconds_map = {
        "1m": 60,
        "3m": 180,
        "5m": 300,
        "15m": 900,
        "30m": 1800,
        "1h": 3600,
        "2h": 7200,
        "4h": 14400,
        "1d": 86400,
    }
    yahoo_interval_map = {
        "1m": "1m",
        "3m": "5m",
        "5m": "5m",
        "15m": "15m",
        "30m": "30m",
        "1h": "60m",
        "2h": "60m",
        "4h": "60m",
        "1d": "1d",
    }
    yahoo_range_map = {
        "1m": "7d",
        "3m": "30d",
        "5m": "30d",
        "15m": "60d",
        "30m": "60d",
        "1h": "730d",
        "2h": "730d",
        "4h": "730d",
        "1d": "10y",
    }
    if interval not in interval_map:
        raise HTTPException(status_code=400, detail="Unsupported interval")

    def normalize_epoch(value: Optional[int]) -> Optional[int]:
        if value is None:
            return None
        return int(value // 1000) if value > 10**12 else int(value)

    from_sec = normalize_epoch(from_ts)
    to_sec = normalize_epoch(to_ts)
    if from_sec is not None and to_sec is not None and from_sec > to_sec:
        raise HTTPException(status_code=400, detail="Invalid candle window: 'from' must be <= 'to'")

    interval_seconds = interval_seconds_map[interval]

    clean_symbol = symbol.upper().replace("/", "").replace("-", "")

    # Map local chart symbols to Twelve Data symbols.
    symbol_map = {
        "BTCUSDT": ["BTC/USD", "BTC/USDT"],
        "ETHUSDT": ["ETH/USD", "ETH/USDT"],
        "BNBUSDT": ["BNB/USD", "BNB/USDT"],
        "SOLUSDT": ["SOL/USD", "SOL/USDT"],
        "XAUUSD": ["XAU/USD", "XAUUSD", "XAU/USD:FX_IDC"],
        "XAGUSD": ["XAG/USD", "XAGUSD", "XAG/USD:FX_IDC"],
        "EURUSD": ["EUR/USD", "EURUSD", "EUR/USD:FX_IDC"],
        "GBPUSD": ["GBP/USD", "GBPUSD", "GBP/USD:FX_IDC"],
        "USDJPY": ["USD/JPY", "USDJPY", "USD/JPY:FX_IDC"],
        "AUDUSD": ["AUD/USD", "AUDUSD", "AUD/USD:FX_IDC"],
        "NAS100": ["NASDAQ", "NDX", "IXIC"],
        "US30": ["DJI", "US30"],
        "SPX500": ["SPX", "GSPC", "SPX500"],
    }
    market_symbol_candidates = symbol_map.get(clean_symbol, [symbol, clean_symbol])

    yahoo_symbol_map = {
        "XAUUSD": "GC=F",
        "XAGUSD": "SI=F",
        "EURUSD": "EURUSD=X",
        "GBPUSD": "GBPUSD=X",
        "USDJPY": "JPY=X",
        "AUDUSD": "AUDUSD=X",
        "NAS100": "^NDX",
        "US30": "^DJI",
        "SPX500": "^GSPC",
        "BTCUSDT": "BTC-USD",
        "ETHUSDT": "ETH-USD",
        "BNBUSDT": "BNB-USD",
        "SOLUSDT": "SOL-USD",
    }

    def parse_twelve_datetime(value: str) -> datetime:
        # Twelve Data can return either intraday datetime or date-only values.
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        raise ValueError(f"Unsupported Twelve Data datetime format: {value}")

    async def fetch_from_twelve_data() -> List[CandleResponse]:
        if not TWELVE_DATA_API_KEY:
            raise HTTPException(status_code=400, detail="TWELVE_DATA_API_KEY is not configured")

        url = "https://api.twelvedata.com/time_series"
        last_error = "No candle data available"

        async with httpx.AsyncClient(timeout=20.0) as client_http:
            for market_symbol in market_symbol_candidates:
                params = {
                    "symbol": market_symbol,
                    "interval": interval_map[interval],
                    "outputsize": limit,
                    "apikey": TWELVE_DATA_API_KEY,
                    "timezone": "UTC",
                }
                if from_sec is not None:
                    params["start_date"] = datetime.fromtimestamp(from_sec, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
                if to_sec is not None:
                    params["end_date"] = datetime.fromtimestamp(to_sec, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
                response = await client_http.get(url, params=params)
                if response.status_code != 200:
                    last_error = f"Twelve Data request failed for {market_symbol}"
                    continue

                payload = response.json()
                values = payload.get("values")
                if not isinstance(values, list) or len(values) == 0:
                    last_error = payload.get("message", f"No candle data for {market_symbol}")
                    continue

                candles: List[CandleResponse] = []
                for item in reversed(values):
                    dt_raw = item.get("datetime")
                    if not dt_raw:
                        continue

                    try:
                        dt = parse_twelve_datetime(dt_raw)
                    except ValueError:
                        continue

                    ts_sec = int(dt.timestamp())
                    if from_sec is not None and ts_sec < from_sec:
                        continue
                    if to_sec is not None and ts_sec > to_sec:
                        continue

                    candles.append(
                        CandleResponse(
                            time=ts_sec * 1000,
                            open=float(item["open"]),
                            high=float(item["high"]),
                            low=float(item["low"]),
                            close=float(item["close"]),
                            volume=float(item.get("volume") or 0),
                        )
                    )

                if candles:
                    return candles[-limit:]

                last_error = f"No valid candle data returned for {market_symbol}"

        raise HTTPException(status_code=400, detail=last_error)

    async def fetch_from_binance() -> List[CandleResponse]:
        url = "https://api.binance.com/api/v3/klines"
        params: Dict[str, Any] = {"symbol": clean_symbol, "interval": interval, "limit": limit}
        if from_sec is not None:
            params["startTime"] = from_sec * 1000
        if to_sec is not None:
            params["endTime"] = to_sec * 1000

        async with httpx.AsyncClient(timeout=15.0) as client_http:
            response = await client_http.get(
                url,
                params=params,
            )

        if response.status_code != 200:
            raise HTTPException(status_code=400, detail="Failed to fetch candles for symbol")

        raw = response.json()
        candles = [
            CandleResponse(
                time=int(item[0]),
                open=float(item[1]),
                high=float(item[2]),
                low=float(item[3]),
                close=float(item[4]),
                volume=float(item[5]),
            )
            for item in raw
        ]
        if from_sec is not None:
            candles = [c for c in candles if (c.time // 1000) >= from_sec]
        if to_sec is not None:
            candles = [c for c in candles if (c.time // 1000) <= to_sec]
        return candles[-limit:]

    async def fetch_from_yahoo() -> List[CandleResponse]:
        yahoo_symbol = yahoo_symbol_map.get(clean_symbol)
        if not yahoo_symbol:
            raise HTTPException(status_code=400, detail=f"No Yahoo symbol mapping for {clean_symbol}")

        interval_value = yahoo_interval_map[interval]
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{yahoo_symbol}"

        params: Dict[str, Any] = {
            "interval": interval_value,
            "includePrePost": "false",
            "events": "div,splits",
        }

        if from_sec is not None or to_sec is not None:
            effective_to = to_sec or int(datetime.now(timezone.utc).timestamp())
            lookback_bars = max(limit * 3, 900)
            effective_from = from_sec or max(0, effective_to - (interval_seconds * lookback_bars))
            params["period1"] = effective_from
            params["period2"] = max(effective_to, effective_from + interval_seconds)
        else:
            params["range"] = yahoo_range_map[interval]

        async with httpx.AsyncClient(timeout=20.0) as client_http:
            response = await client_http.get(
                url,
                params=params,
            )

        if response.status_code != 200:
            raise HTTPException(status_code=400, detail=f"Yahoo request failed for {yahoo_symbol}")

        payload = response.json()
        chart = payload.get("chart", {})
        error = chart.get("error")
        if error:
            raise HTTPException(status_code=400, detail=f"Yahoo error: {error.get('description', 'Unknown error')}")

        result_list = chart.get("result") or []
        if not result_list:
            raise HTTPException(status_code=400, detail=f"No Yahoo data for {yahoo_symbol}")

        result = result_list[0]
        timestamps = result.get("timestamp") or []
        quote = ((result.get("indicators") or {}).get("quote") or [{}])[0]

        opens = quote.get("open") or []
        highs = quote.get("high") or []
        lows = quote.get("low") or []
        closes = quote.get("close") or []
        volumes = quote.get("volume") or []

        candles: List[CandleResponse] = []
        for idx, ts in enumerate(timestamps):
            if idx >= len(opens) or idx >= len(highs) or idx >= len(lows) or idx >= len(closes):
                continue

            o = opens[idx]
            h = highs[idx]
            l = lows[idx]
            c = closes[idx]
            v = volumes[idx] if idx < len(volumes) else 0

            if o is None or h is None or l is None or c is None:
                continue

            candles.append(
                CandleResponse(
                    time=int(ts) * 1000,
                    open=float(o),
                    high=float(h),
                    low=float(l),
                    close=float(c),
                    volume=float(v or 0),
                )
            )

        if not candles:
            raise HTTPException(status_code=400, detail=f"No valid Yahoo candles for {yahoo_symbol}")

        if from_sec is not None:
            candles = [c for c in candles if (c.time // 1000) >= from_sec]
        if to_sec is not None:
            candles = [c for c in candles if (c.time // 1000) <= to_sec]
        if not candles:
            raise HTTPException(status_code=400, detail=f"No Yahoo candles in requested window for {yahoo_symbol}")

        return candles[-limit:]

    try:
        return await fetch_from_twelve_data()
    except HTTPException as twelve_error:
        crypto_symbols = {"BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT"}
        provider_errors = [f"twelve:{twelve_error.detail}"]

        if clean_symbol in crypto_symbols:
            try:
                return await fetch_from_binance()
            except HTTPException as binance_error:
                provider_errors.append(f"binance:{binance_error.detail}")

        try:
            return await fetch_from_yahoo()
        except HTTPException as yahoo_error:
            provider_errors.append(f"yahoo:{yahoo_error.detail}")
            raise HTTPException(status_code=400, detail=" | ".join(provider_errors))
    except Exception as e:
        logging.error(f"Candle API error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch market candles")


@api_router.websocket("/market/live-candles/ws")
async def market_live_candles_ws(
    websocket: WebSocket,
    symbol: str = "BTCUSDT",
    interval: str = "5m",
):
    """Stream the latest candle updates through backend websocket.

    This keeps live chart updates under backend control and lets frontend
    reuse the same transport for all supported symbols.
    """
    await websocket.accept()

    poll_map = {
        "1m": 2.0,
        "5m": 3.0,
        "15m": 4.0,
        "1h": 6.0,
        "4h": 8.0,
        "1d": 12.0,
    }
    poll_seconds = poll_map.get(interval, 3.0)

    last_signature = None

    try:
        while True:
            try:
                candles = await get_market_candles(
                    symbol=symbol,
                    interval=interval,
                    limit=2,
                    from_ts=None,
                    to_ts=None,
                )
            except HTTPException as e:
                await websocket.send_json({"type": "error", "detail": str(e.detail)})
                await asyncio.sleep(poll_seconds)
                continue

            if candles:
                latest = candles[-1]
                signature = (latest.time, latest.close)
                if signature != last_signature:
                    await websocket.send_json({"type": "candle", "data": latest.model_dump()})
                    last_signature = signature

            await asyncio.sleep(poll_seconds)
    except WebSocketDisconnect:
        return
    except Exception as e:
        logging.error(f"Live candle websocket error: {str(e)}")
        try:
            await websocket.close()
        except Exception:
            pass

# ============ AI INSIGHTS ============

@api_router.post("/ai/insights")
async def get_ai_insights(request: AIInsightRequest, current_user: dict = Depends(get_current_user)):
    """Get AI-powered trading insights using xAI Grok (fallback to Groq if needed)."""
    # Log available AI services
    logging.info(f"AI Services - xAI enabled: {xai_grok_enabled}, Groq available: {groq_client is not None}")
    
    # Check if any AI service is available
    if not xai_grok_enabled and not groq_client:
        logging.error("No AI service configured - XAI_API_KEY and GROQ_API_KEY both missing")
        raise HTTPException(status_code=503, detail="AI service is not configured. Please contact support.")
    
    try:
        # Fetch user's closed trades
        trades = await db.trades.find(
            {"user_id": current_user['id']}, 
            {"_id": 0}
        ).sort("created_at", -1).to_list(None)
        
        if not trades:
            return {
                "insight": "No trades found. Start logging your trades to get AI-powered insights.",
                "summary": {
                    "total_trades": 0,
                    "total_pnl": 0,
                    "win_rate": 0,
                    "profit_factor": 0,
                    "expectancy": 0,
                }
            }
        
        # Calculate closed trades and metrics
        closed_trades = [calculate_pnl(t) for t in trades if t['status'] == 'closed']
        
        if not closed_trades:
            return {
                "insight": "You only have open trades. Close some trades first to get meaningful insights.",
                "summary": {
                    "total_trades": 0,
                    "total_pnl": 0,
                    "win_rate": 0,
                    "profit_factor": 0,
                    "expectancy": 0,
                }
            }
        
        # Build trading summary
        total_trades = len(closed_trades)
        total_pnl = sum((t.get('pnl', 0) or 0) for t in closed_trades)
        
        wins = [t for t in closed_trades if (t.get('pnl') or 0) > 0]
        losses = [t for t in closed_trades if (t.get('pnl') or 0) < 0]
        winning = len(wins)
        losing = len(losses)
        win_rate = (winning / total_trades * 100) if total_trades else 0.0
        
        avg_win = (sum((t.get('pnl') or 0) for t in wins) / winning) if winning else 0.0
        avg_loss = (abs(sum((t.get('pnl') or 0) for t in losses)) / losing) if losing else 0.0
        
        gross_profit = sum((t.get('pnl') or 0) for t in wins)
        gross_loss = abs(sum((t.get('pnl') or 0) for t in losses))
        profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else (999 if gross_profit > 0 else 0)
        
        expectancy = (total_pnl / total_trades) if total_trades else 0.0
        
        # Get top/worst instruments
        instrument_stats = {}
        for t in closed_trades:
            inst = t.get('instrument', 'Unknown') or 'Unknown'
            if inst not in instrument_stats:
                instrument_stats[inst] = {"trades": 0, "pnl": 0.0, "wins": 0}
            instrument_stats[inst]["trades"] += 1
            instrument_stats[inst]["pnl"] += t.get('pnl') or 0
            if (t.get('pnl') or 0) > 0:
                instrument_stats[inst]["wins"] += 1
        
        sorted_instruments = sorted(
            instrument_stats.items(),
            key=lambda x: x[1]["pnl"],
            reverse=True
        )
        
        # Build prompt for Groq (Hinglish output)
        user_question = request.question or "Analyze my trading performance and provide insights on how I can improve."
        
        # Cache key based on summary data + question
        cache_key = hashlib.md5(f"{total_trades}:{total_pnl:.2f}:{win_rate:.1f}:{user_question}".encode()).hexdigest()
        cached = ai_insights_cache.get(cache_key)
        if cached and (time.time() - cached['timestamp']) < AI_CACHE_TTL:
            logging.info("AI insights cache hit")
            return cached['response']
        
        summary_text = f"""
Trading Performance Summary:
- Total closed trades: {total_trades}
- Total P&L: ${total_pnl:.2f}
- Win rate: {win_rate:.1f}%
- Average winning trade: ${avg_win:.2f}
- Average losing trade: ${avg_loss:.2f}
- Profit factor: {profit_factor:.2f}
- Expectancy per trade: ${expectancy:.2f}
- Best trade: ${max([t.get('pnl', 0) or 0 for t in closed_trades], default=0):.2f}
- Worst trade: ${min([t.get('pnl', 0) or 0 for t in closed_trades], default=0):.2f}

Top instruments by P&L:
{chr(10).join([f"  {inst}: ${stats['pnl']:.2f} ({stats['wins']}/{stats['trades']} wins)" for inst, stats in sorted_instruments[:3]])}

Recent 10 trades P&L: ${sum(t.get('pnl', 0) or 0 for t in closed_trades[:10]):.2f}

User question: {user_question}
"""
        
        # System prompt for Hinglish responses
        system_prompt = """You are an expert trading coach and performance analyst. You MUST respond in Hinglish (Hindi + English mix), like a trader casually explaining to a friend. Keep it natural, not robotic translation. Use trading terms in English but explain in Hinglish.

Rules:
1. ALWAYS answer the user's EXACT question — do NOT give a generic overview unless asked for one.
2. Use the real trading data provided. Reference specific numbers, instruments, and percentages.
3. Never give short or vague answers. Minimum 4 detailed paragraphs.
4. Use **bold** for key values and section headings.
5. If the user asks about a specific topic (e.g. risk management, mistakes, instruments), focus 80% of your answer on that topic.
6. Include actionable, specific recommendations based on the data.
7. If the user sends a casual message like 'hi' or 'hello', respond naturally but still share one interesting insight from their data.
8. IMPORTANT: Write the ENTIRE response in Hinglish. Mix Hindi and English naturally. Example: "Bhai tera win rate **67.5%** hai jo decent hai, lekin tera risk-reward ratio improve karna padega."
9. Keep instrument names, numbers, and trading terms (stop-loss, take-profit, risk-reward etc.) in English."""

        full_prompt = f"""{summary_text}

All instruments breakdown:
{chr(10).join([f'  {inst}: ${stats["pnl"]:.2f} ({stats["wins"]}/{stats["trades"]} wins, {(stats["wins"]/stats["trades"]*100):.1f}% win rate)' for inst, stats in sorted_instruments])}

Answer this question thoroughly: {user_question}"""
        
        # Try xAI Grok first, fallback to Groq if needed
        insight_text = None
        ai_source = "xai-grok"
        
        try:
            if xai_grok_enabled:
                try:
                    logging.info("Attempting xAI Grok API...")
                    insight_text = await generate_ai_insights_with_xai_grok(full_prompt, system_prompt)
                    logging.info("xAI Grok success")
                except Exception as xai_error:
                    logging.error(f"xAI Grok failed: {str(xai_error)}")
                    if groq_client:
                        logging.info("Falling back to Groq...")
                        ai_source = "groq"
                        groq_start = time.time()
                        message = groq_client.chat.completions.create(
                            model="llama-3.3-70b-versatile",
                            max_tokens=4096,
                            temperature=0.7,
                            messages=[
                                {"role": "system", "content": system_prompt},
                                {"role": "user", "content": full_prompt}
                            ]
                        )
                        insight_text = message.choices[0].message.content
                        groq_elapsed = time.time() - groq_start
                        logging.info(f"Groq fallback success: {groq_elapsed:.2f}s")
                    else:
                        raise xai_error
            elif groq_client:
                # Use Groq directly if xAI not configured
                logging.info("Using Groq API (xAI not configured)...")
                ai_source = "groq"
                groq_start = time.time()
                message = groq_client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    max_tokens=4096,
                    temperature=0.7,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": full_prompt}
                    ]
                )
                insight_text = message.choices[0].message.content
                groq_elapsed = time.time() - groq_start
                logging.info(f"Groq success: {groq_elapsed:.2f}s")
            else:
                raise Exception("No AI service available")
        except Exception as ai_error:
            error_msg = str(ai_error)
            logging.error(f"AI generation error: {error_msg}", exc_info=True)
            
            # Provide specific error messages based on error type
            if "XAI_API_KEY" in error_msg or "xAI Grok API not configured" in error_msg:
                raise HTTPException(
                    status_code=503, 
                    detail=f"xAI Grok API error: {error_msg[:200]}"
                )
            elif "GROQ_API_KEY" in error_msg or "No AI service available" in error_msg:
                raise HTTPException(
                    status_code=503,
                    detail=f"No AI service available: {error_msg[:200]}"
                )
            elif "timeout" in error_msg.lower():
                raise HTTPException(
                    status_code=504,
                    detail=f"AI service timeout: {error_msg[:200]}"
                )
            elif "401" in error_msg or "403" in error_msg or "authentication" in error_msg.lower() or "unauthorized" in error_msg.lower():
                raise HTTPException(
                    status_code=503,
                    detail=f"AI API authentication failed: {error_msg[:200]}. Please check your API keys."
                )
            elif "Grok API failed" in error_msg:
                # Extract the actual xAI error
                raise HTTPException(
                    status_code=503,
                    detail=f"xAI Grok API error: {error_msg[:300]}"
                )
            else:
                raise HTTPException(
                    status_code=500, 
                    detail=f"AI generation failed: {error_msg[:200]}"
                )
        
        if not insight_text:
            raise HTTPException(status_code=500, detail="Failed to generate insights")
        
        # Generate TTS audio from Hinglish text
        audio_base64 = None
        try:
            audio_bytes = await generate_tts(insight_text)
            if audio_bytes:
                audio_base64 = base64.b64encode(audio_bytes).decode("utf-8")
        except Exception as tts_err:
            logging.error(f"TTS generation failed (non-blocking): {str(tts_err)}")
        
        # Build chart data for frontend visualization
        instrument_chart = [
            {
                "name": inst,
                "pnl": round(stats["pnl"], 2),
                "trades": stats["trades"],
                "wins": stats["wins"],
                "winRate": round((stats["wins"] / stats["trades"] * 100), 1) if stats["trades"] > 0 else 0,
            }
            for inst, stats in sorted_instruments
        ]
        
        # Direction breakdown (long vs short)
        long_trades = [t for t in closed_trades if (t.get('direction') or '').lower() in ('long', 'buy')]
        short_trades = [t for t in closed_trades if (t.get('direction') or '').lower() in ('short', 'sell')]
        direction_chart = {
            "long": {
                "total": len(long_trades),
                "wins": len([t for t in long_trades if (t.get('pnl') or 0) > 0]),
                "pnl": round(sum(t.get('pnl', 0) or 0 for t in long_trades), 2),
            },
            "short": {
                "total": len(short_trades),
                "wins": len([t for t in short_trades if (t.get('pnl') or 0) > 0]),
                "pnl": round(sum(t.get('pnl', 0) or 0 for t in short_trades), 2),
            },
        }
        
        best_trade = max(closed_trades, key=lambda t: t.get('pnl', 0) or 0, default={})
        worst_trade = min(closed_trades, key=lambda t: t.get('pnl', 0) or 0, default={})
        
        # Detect currency from trades
        ai_currency = 'USD'
        for t in closed_trades:
            if t.get('currency'):
                ai_currency = t['currency']
                break
        
        response_data = {
            "insight": insight_text,
            "audio": audio_base64,
            "summary": {
                "total_trades": total_trades,
                "total_pnl": round(total_pnl, 2),
                "win_rate": round(win_rate, 2),
                "profit_factor": round(profit_factor, 2),
                "expectancy": round(expectancy, 2),
                "avg_win": round(avg_win, 2),
                "avg_loss": round(avg_loss, 2),
                "winning_trades": winning,
                "losing_trades": losing,
                "gross_profit": round(gross_profit, 2),
                "gross_loss": round(gross_loss, 2),
                "currency": ai_currency,
            },
            "charts": {
                "instruments": instrument_chart,
                "direction": direction_chart,
                "best_trade": {
                    "pnl": round(best_trade.get('pnl', 0) or 0, 2),
                    "instrument": best_trade.get('instrument', 'N/A'),
                },
                "worst_trade": {
                    "pnl": round(worst_trade.get('pnl', 0) or 0, 2),
                    "instrument": worst_trade.get('instrument', 'N/A'),
                },
            },
            "source": ai_source
        }
        # Cache the response
        ai_insights_cache[cache_key] = {'response': response_data, 'timestamp': time.time()}
        if len(ai_insights_cache) > 50:
            oldest_key = min(ai_insights_cache, key=lambda k: ai_insights_cache[k]['timestamp'])
            del ai_insights_cache[oldest_key]
        return response_data
    except Exception as e:
        logging.error(f"AI insights error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate insights: {str(e)}")


# ============ IMPORT ROUTES ============

# Symbol mapping for CSV import (broker symbols to app instruments)
SYMBOL_MAPPING = {
    'XAUUSDm': 'XAU/USD',
    'XAUUSD': 'XAU/USD',
    'XAUUSD+': 'XAU/USD',
    'XAGUSDm': 'XAG/USD',
    'XAGUSD': 'XAG/USD',
    'XAGUSD+': 'XAG/USD',
    'BTCUSDm': 'BTC/USD',
    'BTCUSD': 'BTC/USD',
    'BTCUSD+': 'BTC/USD',
    'ETHUSDm': 'ETH/USD',
    'ETHUSD': 'ETH/USD',
    'ETHUSD+': 'ETH/USD',
    'EURUSDm': 'EUR/USD',
    'EURUSD': 'EUR/USD',
    'EURUSD+': 'EUR/USD',
    'GBPUSDm': 'GBP/USD',
    'GBPUSD': 'GBP/USD',
    'GBPUSD+': 'GBP/USD',
    'USDJPYm': 'USD/JPY',
    'USDJPY': 'USD/JPY',
    'USDJPY+': 'USD/JPY',
    'NAS100m': 'NAS100',
    'NAS100': 'NAS100',
    'US30m': 'US30',
    'US30': 'US30',
}

def normalize_symbol(symbol: str) -> str:
    """Convert broker symbol format to app instrument format"""
    symbol = symbol.strip()
    if not symbol:
        return symbol
    # Normalize case so broker variants like xauusdm or XauUsdM map reliably.
    symbol = symbol.upper()
    if symbol in SYMBOL_MAPPING:
        return SYMBOL_MAPPING[symbol]
    # Remove trailing 'm' or '+' if present and check again
    for suffix in ['M', '+', '.', '#']:
        if symbol.endswith(suffix):
            base = symbol[:-1]
            if base in SYMBOL_MAPPING:
                return SYMBOL_MAPPING[base]
    # Return original if no mapping found
    return symbol

class CSVImportResponse(BaseModel):
    success: bool
    imported_count: int
    skipped_count: int
    errors: List[str] = []
    message: str

@api_router.post("/trades/import-csv", response_model=CSVImportResponse)
async def import_trades_csv(
    file: UploadFile = File(...),
    force: bool = Query(default=False),
    current_user: dict = Depends(get_current_user)
):
    """Import trades from a CSV file (supports MT5 and app-export format)."""

    file_name = (file.filename or '').lower()
    if not file_name.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are accepted")
    
    try:
        # Read and decode the CSV content
        content = await file.read()
        try:
            decoded = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded = content.decode('latin-1')
        
        # Parse CSV
        reader = csv.DictReader(io.StringIO(decoded))
        raw_headers = [h for h in (reader.fieldnames or []) if h]

        def normalize_header_key(value: str) -> str:
            if value is None:
                return ''
            return ''.join(ch for ch in str(value).strip().lower() if ch.isalnum())

        normalized_headers = {normalize_header_key(h) for h in raw_headers if normalize_header_key(h)}

        if not normalized_headers:
            raise HTTPException(status_code=400, detail="CSV file is empty or missing header row")

        def get_row_value(row_lookup: Dict[str, Any], keys: List[str], default: str = "") -> str:
            for key in keys:
                normalized_key = normalize_header_key(key)
                value = row_lookup.get(normalized_key)
                if value is not None and str(value).strip() != "":
                    return str(value).strip()
            return default

        def parse_float(value: str, field_label: str, row_num: int, required: bool = False) -> Optional[float]:
            if value is None or value == "":
                if required:
                    raise ValueError(f"Row {row_num}: Missing {field_label}")
                return None
            try:
                return float(value)
            except ValueError:
                raise ValueError(f"Row {row_num}: Invalid {field_label}")

        def normalize_date(value: str) -> Optional[str]:
            if not value:
                return None
            # Keep a compact, stable date-time string and support date-only input.
            return value.replace(' ', 'T')[:19]

        def build_trade_fingerprint(trade: dict) -> str:
            return "|".join([
                str(trade.get("instrument") or ""),
                str(trade.get("position") or ""),
                str(round(float(trade.get("entry_price") or 0), 8)),
                str(round(float(trade.get("exit_price") or 0), 8)) if trade.get("exit_price") is not None else "",
                str(round(float(trade.get("quantity") or 0), 8)),
                str(trade.get("entry_date") or ""),
                str(trade.get("exit_date") or ""),
                str(trade.get("status") or ""),
            ])

        def build_mt5_signature(trade: dict) -> str:
            return "|".join([
                str(trade.get("mt5_ticket") or ""),
                str(trade.get("entry_date") or ""),
                str(trade.get("exit_date") or ""),
                str(round(float(trade.get("quantity") or 0), 8)),
                str(round(float(trade.get("entry_price") or 0), 8)),
                str(round(float(trade.get("exit_price") or 0), 8)) if trade.get("exit_price") is not None else "",
            ])
        
        imported_count = 0
        skipped_count = 0
        duplicate_count = 0
        error_count = 0
        errors = []

        existing_trade_count = await db.trades.count_documents({"user_id": current_user['id']})
        
        # Use a row-level MT5 signature so partial closes with same ticket can still import.
        existing_mt5_signatures = set()
        if existing_trade_count > 0:
            existing_trades = await db.trades.find(
                {"user_id": current_user['id'], "mt5_ticket": {"$ne": None}},
                {
                    "_id": 0,
                    "mt5_ticket": 1,
                    "entry_date": 1,
                    "exit_date": 1,
                    "quantity": 1,
                    "entry_price": 1,
                    "exit_price": 1,
                }
            ).to_list(100000)
            for t in existing_trades:
                if t.get('mt5_ticket'):
                    existing_mt5_signatures.add(build_mt5_signature(t))

        # Fallback duplicate detection for files without ticket numbers.
        existing_fingerprints = set()
        if existing_trade_count > 0:
            existing_for_fingerprint = await db.trades.find(
                {"user_id": current_user['id']},
                {
                    "_id": 0,
                    "instrument": 1,
                    "position": 1,
                    "entry_price": 1,
                    "exit_price": 1,
                    "quantity": 1,
                    "entry_date": 1,
                    "exit_date": 1,
                    "status": 1,
                }
            ).to_list(100000)
            for existing in existing_for_fingerprint:
                existing_fingerprints.add(build_trade_fingerprint(existing))
        
        now = datetime.now(timezone.utc).isoformat()
        trades_to_insert = []

        has_mt5_shape = {'ticket', 'symbol', 'type', 'openingprice', 'lots'}.issubset(normalized_headers)
        has_export_shape = {'instrument', 'position', 'entryprice', 'quantity', 'entrydate'}.issubset(normalized_headers)

        if not has_mt5_shape and not has_export_shape:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Unsupported CSV format. Expected MT5 columns (ticket,symbol,type,opening_price,lots,...) "
                    "or app export columns (Instrument,Position,Entry Price,Quantity,Entry Date,...)."
                )
            )
        
        for row_num, row in enumerate(reader, start=2):
            try:
                row_lookup = {
                    normalize_header_key(key): value
                    for key, value in (row or {}).items()
                    if key is not None
                }

                ticket = get_row_value(row_lookup, ['ticket'])
                
                # Parse shared fields from either MT5 or app-export headers.
                symbol = get_row_value(row_lookup, ['symbol', 'instrument'])
                if not symbol:
                    errors.append(f"Row {row_num}: Missing symbol")
                    error_count += 1
                    skipped_count += 1
                    continue
                
                raw_trade_type = get_row_value(row_lookup, ['type', 'position']).lower()
                trade_type = {'long': 'buy', 'short': 'sell'}.get(raw_trade_type, raw_trade_type)
                if trade_type not in ['buy', 'sell']:
                    errors.append(f"Row {row_num}: Invalid trade type '{trade_type}'")
                    error_count += 1
                    skipped_count += 1
                    continue
                
                # Parse prices
                try:
                    opening_price = parse_float(
                        get_row_value(row_lookup, ['opening_price', 'entry_price', 'Entry Price']),
                        'entry/opening price',
                        row_num,
                        required=True,
                    )
                    closing_price = parse_float(
                        get_row_value(row_lookup, ['closing_price', 'exit_price', 'Exit Price']),
                        'exit/closing price',
                        row_num,
                        required=False,
                    )
                except ValueError as e:
                    errors.append(str(e))
                    error_count += 1
                    skipped_count += 1
                    continue
                
                # Parse lots
                try:
                    lots = parse_float(
                        get_row_value(row_lookup, ['lots', 'quantity', 'Quantity']),
                        'lot size/quantity',
                        row_num,
                        required=True,
                    )
                    if lots is None or lots <= 0:
                        raise ValueError(f"Row {row_num}: Invalid lot size")
                except ValueError as e:
                    errors.append(str(e))
                    error_count += 1
                    skipped_count += 1
                    continue
                
                # Parse dates
                opening_time = get_row_value(row_lookup, ['opening_time_utc', 'entry_date'])
                closing_time = get_row_value(row_lookup, ['closing_time_utc', 'exit_date'])
                
                # Convert datetime format (2026-03-02T01:33:34.811 -> 2026-03-02T01:33:34)
                entry_date = normalize_date(opening_time) or now[:19]
                exit_date = normalize_date(closing_time)
                
                # Parse optional fields
                stop_loss = None
                take_profit = None
                commission = 0
                swap = 0
                profit = None
                note_text = get_row_value(row_lookup, ['notes'])
                explicit_status = get_row_value(row_lookup, ['status']).lower()
                
                try:
                    stop_loss_value = get_row_value(row_lookup, ['stop_loss'])
                    take_profit_value = get_row_value(row_lookup, ['take_profit'])
                    commission_value = get_row_value(row_lookup, ['commission_usd', 'commission'], '0')
                    swap_value = get_row_value(row_lookup, ['swap_usd', 'swap'], '0')
                    profit_value = get_row_value(row_lookup, ['profit_usd', 'pnl', 'p&l', 'pl'])

                    if stop_loss_value:
                        stop_loss = float(stop_loss_value)
                    if take_profit_value:
                        take_profit = float(take_profit_value)
                    if commission_value:
                        commission = float(commission_value)
                    if swap_value:
                        swap = float(swap_value)
                    if profit_value:
                        profit = float(profit_value)
                except ValueError:
                    pass  # Use defaults for optional fields
                
                # Normalize symbol to app format
                instrument = normalize_symbol(symbol)
                
                # Determine status
                if explicit_status in {'open', 'closed'}:
                    status = explicit_status
                else:
                    status = 'closed' if closing_price is not None and exit_date else 'open'
                
                # Create trade document
                trade_id = str(uuid.uuid4())
                trade_doc = {
                    "id": trade_id,
                    "user_id": current_user['id'],
                    "instrument": instrument,
                    "position": trade_type,
                    "entry_price": opening_price,
                    "exit_price": closing_price,
                    "quantity": lots,
                    "entry_date": entry_date,
                    "exit_date": exit_date,
                    "notes": note_text or (f"Imported from CSV - Ticket #{ticket}" if ticket else "Imported from CSV"),
                    "status": status,
                    "stop_loss": stop_loss,
                    "take_profit": take_profit,
                    "commission": commission,
                    "swap": swap,
                    "mt5_ticket": ticket or None,
                    "created_at": now
                }

                # Skip duplicates for non-ticket imports.
                fingerprint = build_trade_fingerprint(trade_doc)
                if not force and not ticket and fingerprint in existing_fingerprints:
                    duplicate_count += 1
                    skipped_count += 1
                    continue

                if not force and ticket:
                    mt5_signature = build_mt5_signature(trade_doc)
                    if mt5_signature in existing_mt5_signatures:
                        duplicate_count += 1
                        skipped_count += 1
                        continue
                
                # Use the profit from CSV directly if available, otherwise calculate
                if profit is not None and status == 'closed':
                    trade_doc['pnl'] = round(profit, 2)
                    # Calculate pnl_percentage based on entry price
                    if opening_price > 0 and closing_price is not None:
                        if trade_type == 'buy':
                            trade_doc['pnl_percentage'] = round((closing_price - opening_price) / opening_price * 100, 2)
                        else:
                            trade_doc['pnl_percentage'] = round((opening_price - closing_price) / opening_price * 100, 2)
                else:
                    trade_doc = calculate_pnl(trade_doc)
                
                trades_to_insert.append(trade_doc)
                if ticket:
                    existing_mt5_signatures.add(build_mt5_signature(trade_doc))
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
                skipped_count += 1
        
        # Bulk insert trades
        if trades_to_insert:
            await db.trades.insert_many(trades_to_insert)
        
        # Limit errors list for response
        if len(errors) > 10:
            errors = errors[:10] + [f"... and {len(errors) - 10} more errors"]
        
        if imported_count > 0:
            message = (
                f"Successfully imported {imported_count} trades. "
                f"{duplicate_count} duplicates skipped, {error_count} rows had errors."
            )
        elif skipped_count > 0 and error_count == 0:
            message = "No new trades imported. All rows are duplicates of existing entries."
        elif skipped_count > 0:
            message = (
                f"No new trades imported. {error_count} rows had errors "
                f"and {duplicate_count} rows were duplicates."
            )
        else:
            message = "No trade rows found in the uploaded CSV."

        return CSVImportResponse(
            success=True,
            imported_count=imported_count,
            skipped_count=skipped_count,
            errors=errors,
            message=message
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"CSV import error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to import CSV: {str(e)}")

# ============ EXPORT ROUTES ============

@api_router.get("/export/trades/csv")
async def export_trades_csv(current_user: dict = Depends(get_current_user)):
    trades = await db.trades.find({"user_id": current_user['id']}, {"_id": 0}).sort("entry_date", -1).to_list(10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'Instrument', 'Position', 'Status', 'Entry Price', 'Exit Price', 
        'Quantity', 'Entry Date', 'Exit Date', 'P&L', 'P&L %',
        'Stop Loss', 'Take Profit', 'Commission', 'Swap', 'Notes'
    ])
    
    for trade in trades:
        trade = calculate_pnl(trade)
        writer.writerow([
            trade.get('instrument', ''),
            trade.get('position', ''),
            trade.get('status', ''),
            trade.get('entry_price', ''),
            trade.get('exit_price', ''),
            trade.get('quantity', ''),
            trade.get('entry_date', ''),
            trade.get('exit_date', ''),
            trade.get('pnl', ''),
            trade.get('pnl_percentage', ''),
            trade.get('stop_loss', ''),
            trade.get('take_profit', ''),
            trade.get('commission', ''),
            trade.get('swap', ''),
            trade.get('notes', '')
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=trades_export_{datetime.now().strftime('%Y%m%d')}.csv"}
    )

@api_router.get("/export/trades/xlsx")
async def export_trades_xlsx(current_user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    trades = await db.trades.find({"user_id": current_user['id']}, {"_id": 0}).sort("entry_date", -1).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = [
        'Instrument', 'Position', 'Status', 'Entry Price', 'Exit Price', 
        'Quantity', 'Entry Date', 'Exit Date', 'P&L', 'P&L %',
        'Stop Loss', 'Take Profit', 'Commission', 'Swap', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    for row_idx, trade in enumerate(trades, 2):
        trade = calculate_pnl(trade)
        row_data = [
            trade.get('instrument', ''),
            trade.get('position', ''),
            trade.get('status', ''),
            trade.get('entry_price', ''),
            trade.get('exit_price', ''),
            trade.get('quantity', ''),
            trade.get('entry_date', ''),
            trade.get('exit_date', ''),
            trade.get('pnl', ''),
            trade.get('pnl_percentage', ''),
            trade.get('stop_loss', ''),
            trade.get('take_profit', ''),
            trade.get('commission', ''),
            trade.get('swap', ''),
            trade.get('notes', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            
            if col == 9 and value:
                try:
                    if float(value) > 0:
                        cell.fill = green_fill
                    elif float(value) < 0:
                        cell.fill = red_fill
                except:
                    pass
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=trades_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# ============ HEALTH CHECK ============

@api_router.get("/")
async def root():
    return {"message": "TradeLedger API", "status": "healthy"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Forex News endpoint
@api_router.get("/news")
async def get_forex_news():
    """Fetch latest forex news from Finnhub API with caching"""
    global news_cache
    
    # Check if cache is valid
    if news_cache['data'] and news_cache['timestamp']:
        elapsed = (datetime.now(timezone.utc) - news_cache['timestamp']).total_seconds()
        if elapsed < NEWS_CACHE_TTL:
            return news_cache['data']
    
    # Fetch fresh news
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(
                f"https://finnhub.io/api/v1/news?category=forex&token={FINNHUB_API_KEY}",
                timeout=10.0
            )
            response.raise_for_status()
            news_data = response.json()
        
        # Filter and format news
        forex_keywords = [
            "USD", "EUR", "GBP", "JPY", "XAU", "Gold", "Fed", "ECB", "BOJ", "BOE",
            "Inflation", "Interest Rate", "CPI", "NFP", "Dollar", "Forex", "FX",
            "Central Bank", "Rate Hike", "Rate Cut", "FOMC", "Monetary Policy"
        ]
        
        filtered_news = []
        for item in news_data[:20]:  # Process more to get enough filtered results
            headline = item.get('headline', '')
            
            # Check if headline contains forex keywords
            if any(keyword.lower() in headline.lower() for keyword in forex_keywords):
                # Calculate relative time
                news_time = datetime.fromtimestamp(item.get('datetime', 0), tz=timezone.utc)
                now = datetime.now(timezone.utc)
                diff = now - news_time
                
                if diff.days > 0:
                    time_str = f"{diff.days}d ago"
                elif diff.seconds >= 3600:
                    time_str = f"{diff.seconds // 3600}h ago"
                elif diff.seconds >= 60:
                    time_str = f"{diff.seconds // 60}m ago"
                else:
                    time_str = "Just now"
                
                # Check for high-impact keywords
                high_impact = any(word.lower() in headline.lower() 
                                for word in ["Fed", "Interest Rate", "CPI", "NFP", "FOMC", "Rate Hike", "Rate Cut"])
                
                filtered_news.append({
                    "headline": headline,
                    "source": item.get('source', 'Unknown'),
                    "time": time_str,
                    "url": item.get('url', ''),
                    "highImpact": high_impact
                })
                
                if len(filtered_news) >= 15:
                    break
        
        # Update cache
        news_cache['data'] = filtered_news
        news_cache['timestamp'] = datetime.now(timezone.utc)
        
        return filtered_news
        
    except Exception as e:
        logging.error(f"Failed to fetch forex news: {str(e)}")
        # Return cached data if available, even if expired
        if news_cache['data']:
            return news_cache['data']
        raise HTTPException(status_code=503, detail="Failed to fetch forex news")


# Economic Calendar endpoint
@api_router.get("/calendar")
async def get_economic_calendar():
    """Fetch economic events for today from Finnhub API with caching"""
    global calendar_cache
    
    # Check if cache is valid
    if calendar_cache['data'] and calendar_cache['timestamp']:
        elapsed = (datetime.now(timezone.utc) - calendar_cache['timestamp']).total_seconds()
        if elapsed < CALENDAR_CACHE_TTL:
            print(f"Returning cached calendar data ({len(calendar_cache['data'])} events)")
            return calendar_cache['data']
    
    # Fetch fresh calendar data
    today = date.today()
    url = f"https://finnhub.io/api/v1/calendar/economic?from={today}&to={today}&token={FINNHUB_API_KEY}"
    
    try:
        print(f"\n{'='*60}")
        print(f"🔥 STEP 1: Calling Finnhub Calendar API")
        print(f"{'='*60}")
        print(f"Date: {today}")
        print(f"URL: {url[:100]}...")
        
        async with httpx.AsyncClient() as client:
            response = await client.get(url, timeout=10.0)
        
        print(f"\n🔥 STEP 2: Response received")
        print(f"Status Code: {response.status_code}")
        
        if response.status_code != 200:
            print(f"❌ Non-200 status code: {response.status_code}")
            print(f"Response text: {response.text[:500]}")
            raise Exception(f"API returned status {response.status_code}")
        
        data = response.json()
        print(f"Response Type: {type(data)}")
        print(f"Response Keys: {list(data.keys()) if isinstance(data, dict) else 'Not a dict'}")
        
        # Check for API errors
        if isinstance(data, dict) and 'error' in data:
            print(f"❌ API Error: {data.get('error')}")
            raise Exception(f"API error: {data.get('error')}")
        
        events = data.get("economicCalendar", [])
        print(f"\n🔥 STEP 3: Events extracted")
        print(f"Total events in response: {len(events)}")
        
        if not events:
            print("⚠️ No events returned from API (this is NORMAL, not an error)")
            fallback = [{
                "time": "All Day",
                "currency": "GLOBAL",
                "event": "No major economic events scheduled for today",
                "impact": "low"
            }]
            calendar_cache['data'] = fallback
            calendar_cache['timestamp'] = datetime.now(timezone.utc)
            return fallback
        
        # Process events (show first 3 for debugging)
        filtered = []
        for idx, e in enumerate(events[:15]):
            if idx < 3:
                print(f"\nEvent {idx + 1} sample:")
                print(f"  Time: {e.get('time', 'N/A')}")
                print(f"  Country: {e.get('country', 'N/A')}")
                print(f"  Currency: {e.get('currency', 'N/A')}")
                print(f"  Event: {e.get('event', 'N/A')}")
                print(f"  Impact: {e.get('impact', 'N/A')}")
            
            filtered.append({
                "time": e.get("time", "TBA"),
                "currency": e.get("country") or e.get("currency") or "N/A",
                "event": e.get("event", "Economic Event"),
                "impact": e.get("impact", "N/A")
            })
        
        print(f"\n🔥 STEP 4: Filtered events")
        print(f"Total filtered: {len(filtered)}")
        
        # Update cache
        calendar_cache['data'] = filtered
        calendar_cache['timestamp'] = datetime.now(timezone.utc)
        
        print(f"✅ SUCCESS: Returning {len(filtered)} events")
        print(f"{'='*60}\n")
        
        return filtered
        
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ ERROR in fetch_calendar")
        print(f"{'='*60}")
        print(f"Error Type: {type(e).__name__}")
        print(f"Error Message: {str(e)}")
        print(f"{'='*60}\n")
        
        # Return cached data if available (even if expired)
        if calendar_cache['data']:
            print(f"⚠️ Returning cached data instead ({len(calendar_cache['data'])} events)")
            return calendar_cache['data']
        
        # Return user-friendly fallback (NOT an error message)
        print("⚠️ No cache available - returning fallback message")
        return [{
            "time": "N/A",
            "currency": "SYSTEM",
            "event": "Calendar temporarily unavailable - please refresh",
            "impact": "low"
        }]


# Market Quotes endpoint - Hybrid approach using Finnhub + Live-Rates.com
@api_router.get("/quotes")
async def get_market_quotes():
    """Fetch live market quotes using Finnhub (crypto/indices) + Live-Rates.com (metals/oil)"""
    global quotes_cache
    
    # Check if cache is valid
    if quotes_cache['data'] and quotes_cache['timestamp']:
        elapsed = (datetime.now(timezone.utc) - quotes_cache['timestamp']).total_seconds()
        if elapsed < QUOTES_CACHE_TTL:
            return quotes_cache['data']
    
    try:
        print(f"\n{'='*60}")
        print(f"📊 Fetching market quotes from multiple sources")
        print(f"{'='*60}")
        
        results = []
        
        async with httpx.AsyncClient() as client:
            # 1. CRYPTO from Finnhub (fast and reliable)
            crypto_symbols = [
                {"symbol": "BINANCE:BTCUSDT", "display": "BTC"},
                {"symbol": "BINANCE:ETHUSDT", "display": "ETH"},
                {"symbol": "BINANCE:SOLUSDT", "display": "SOL"}
            ]
            
            for item in crypto_symbols:
                try:
                    url = f"https://finnhub.io/api/v1/quote?symbol={item['symbol']}&token={FINNHUB_API_KEY}"
                    response = await client.get(url, timeout=5.0)
                    
                    if response.status_code == 200:
                        data = response.json()
                        current_price = data.get('c', 0)
                        change = data.get('d', 0)
                        percent = data.get('dp', 0)
                        
                        if current_price > 0:
                            results.append({
                                "symbol": item['display'],
                                "price": round(current_price, 2),
                                "change": round(change, 2),
                                "percent": round(percent, 2)
                            })
                            print(f"✅ [Finnhub] {item['display']}: ${current_price:,.2f} ({percent:+.2f}%)")
                except Exception as e:
                    print(f"❌ Error fetching {item['display']}: {str(e)}")
            
            # 2. METALS & OIL from Live-Rates.com (FREE, real spot prices!)
            try:
                url = "https://www.live-rates.com/rates"
                response = await client.get(url, timeout=10.0)
                
                if response.status_code == 200:
                    data = response.json()
                    
                    # Map of what we want to extract
                    commodity_map = {
                        "XAU/USD": "GOLD",
                        "XAG/USD": "SILVER",
                        "USOil": "OIL"  # Live-Rates uses "USOil" not "WTI"
                    }
                    
                    for item in data:
                        currency = item.get("currency", "")
                        if currency in commodity_map:
                            price = float(item.get("rate", 0))
                            bid = float(item.get("bid", price))
                            ask = float(item.get("ask", price))
                            open_price = float(item.get("open", price))
                            
                            if price > 0 and open_price > 0:
                                change = price - open_price
                                percent = (change / open_price) * 100
                                
                                results.append({
                                    "symbol": commodity_map[currency],
                                    "price": round(price, 2),
                                    "change": round(change, 2),
                                    "percent": round(percent, 2)
                                })
                                print(f"✅ [Live-Rates] {commodity_map[currency]}: ${price:,.2f} ({percent:+.2f}%)")
                else:
                    print(f"⚠️ Live-Rates returned {response.status_code}")
                    
            except Exception as e:
                print(f"❌ Error fetching commodities from Live-Rates: {str(e)}")
            
            # 3. US INDICES - using major ETFs from Finnhub
            index_symbols = [
                {"symbol": "SPY", "display": "S&P500"},
                {"symbol": "QQQ", "display": "NASDAQ"},
                {"symbol": "DIA", "display": "DOW"}
            ]
            
            for item in index_symbols:
                try:
                    url = f"https://finnhub.io/api/v1/quote?symbol={item['symbol']}&token={FINNHUB_API_KEY}"
                    response = await client.get(url, timeout=5.0)
                    
                    if response.status_code == 200:
                        data = response.json()
                        current_price = data.get('c', 0)
                        change = data.get('d', 0)
                        percent = data.get('dp', 0)
                        
                        if current_price > 0:
                            results.append({
                                "symbol": item['display'],
                                "price": round(current_price, 2),
                                "change": round(change, 2),
                                "percent": round(percent, 2)
                            })
                            print(f"✅ [Finnhub] {item['display']}: ${current_price:,.2f} ({percent:+.2f}%)")
                except Exception as e:
                    print(f"❌ Error fetching {item['display']}: {str(e)}")
        
        # Cache and return results
        if results:
            quotes_cache['data'] = results
            quotes_cache['timestamp'] = datetime.now(timezone.utc)
            print(f"✅ Successfully fetched {len(results)}/9 quotes")
            print(f"{'='*60}\n")
            return results
        else:
            # Return cached data if available
            if quotes_cache['data']:
                print("⚠️ No new data, returning cached quotes")
                return quotes_cache['data']
            
            print("⚠️ No quotes available")
            return []
            
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ ERROR in quotes endpoint")
        print(f"Error: {str(e)}")
        print(f"{'='*60}\n")
        
        # Return cached data if available
        if quotes_cache['data']:
            return quotes_cache['data']
        
        return []


# Include routers
app.include_router(api_router)
app.include_router(admin_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
